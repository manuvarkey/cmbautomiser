#!/usr/bin/env python
# -*- coding: utf-8 -*-
#
# bill.py
#
#  Copyright 2014 Manu Varkey <manuvarkey@gmail.com>
#
#  This program is free software; you can redistribute it and/or modify
#  it under the terms of the GNU General Public License as published by
#  the Free Software Foundation; either version 2 of the License, or
#  (at your option) any later version.
#
#  This program is distributed in the hope that it will be useful,
#  but WITHOUT ANY WARRANTY; without even the implied warranty of
#  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#  GNU General Public License for more details.
#
#  You should have received a copy of the GNU General Public License
#  along with this program; if not, write to the Free Software
#  Foundation, Inc., 51 Franklin Street, Fifth Floor, Boston,
#  MA 02110-1301, USA.
#
#

import pickle
import copy
import math

from gi.repository import Gtk, Gdk, GLib

from undo import *

# local files import
from globalconstants import *
from schedule import *
from cmb import *
from bill_dialog import *
from misc import *

from openpyxl import Workbook, load_workbook, worksheet
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.cell import get_column_letter

# Data structure for Bill Class
class BillData:
    def __init__(self, bill_type=BILL_NORMAL):
        self.prev_bill = None
        self.cmb_name = ''
        self.title = ''
        self.bill_date = ''
        self.starting_page = 1
        self.mitems = []  # for storing paths to billed measurement items [cmb][measurement][Meas Item]
        self.item_part_percentage = []  # part rate for exess rate items
        self.item_excess_part_percentage = []  # part rate for exess rate items
        self.item_excess_rates = []  # list of excess rates above excess_percentage

        # additional elements for custom bill
        self.item_qty = []  # qtys of items b/f
        self.item_normal_amount = []  # total item amount for qty at normal rate
        self.item_excess_amount = []  # amounts for qty at excess rate

        self.bill_type = bill_type


# class storing bill of rates for work
class Bill:
    def __init__(self, bill_view):
        self.bill_view = bill_view
        self.data = BillData()

        # schedule item vars
        self.item_cmb_ref = [[]]  # cmb references of items b/f
        self.item_paths = [[]]
        self.item_qty = [[]]  # qtys of items b/f
        self.item_normal_qty = []  # list of excess qty above excess_percentage
        self.item_normal_amount = []  # total item amount for qty at normal rate
        self.item_excess_qty = []  # qty at excess rate
        self.item_excess_amount = []  # amounts for qty at excess rate

        self.prev_bill = None  # prev_bill data object
        self.cmb_ref = set()  # set containing refered cmbs
        self.bill_total_amount = 0  # total amount of work done uptodate
        self.bill_since_prev_amount = 0  # since previous amount of work done

        self.schedule = bill_view.schedule  # main data store of rows
        self.cmbs = bill_view.cmbs  # list of cmbs

    def clear(self):
        bill_type = self.data.bill_type
        self.data = BillData(bill_type)

        # schedule item vars
        self.item_cmb_ref = [[]]  # cmb references of items b/f
        self.item_paths = [[]]
        self.item_qty = [[]]  # qtys of items b/f
        self.item_normal_qty = []  # list of excess qty above excess_percentage
        self.item_normal_amount = []  # total item amount for qty at normal rate
        self.item_excess_qty = []  # qty at excess rate
        self.item_excess_amount = []  # amounts for qty at excess rate

        self.prev_bill = None  # prev_bill data object
        self.cmb_ref = set()  # set containing refered cmbs
        self.bill_total_amount = 0  # total amount of work done uptodate
        self.bill_since_prev_amount = 0  # since previous amount of work done

    def get_modal(self):
        return self.data

    def set_modal(self, data):
        self.data = data
        for path in self.data.mitems:  # check and remove items already measured
            if self.cmbs[path[0]][path[1]][path[2]].get_billed_flag() == True:
                del (path)
        self.update_values()

    def update_values(self):
        sch_len = self.schedule.length()
        if self.data.bill_type == BILL_NORMAL:
            # Update schedule item vars
            if self.data.prev_bill is not None:
                self.prev_bill = self.bill_view.bills[self.data.prev_bill]  # get prev_bill object
            else:
                self.prev_bill = None
            # initialise variables to schedule length
            self.item_cmb_ref = []
            self.item_paths = []
            self.item_qty = []
            for i in range(sch_len):  # make a list of empty lists
                self.item_cmb_ref.append([])
                self.item_paths.append([])
                self.item_qty.append([])
            self.item_normal_qty = [0] * sch_len
            self.item_normal_amount = [0] * sch_len
            self.item_excess_qty = [0] * sch_len
            self.item_excess_amount = [0] * sch_len

            # expand variables to schedule length
            existing_length = len(self.data.item_excess_rates)
            if sch_len > existing_length:
                self.data.item_part_percentage += [100] * (sch_len - existing_length)
                self.data.item_excess_part_percentage += [100] * (sch_len - existing_length)
                self.data.item_excess_rates += [0] * (sch_len - existing_length)
                for i in range(sch_len - existing_length):
                    self.data.item_qty.append([0])
                self.data.item_normal_amount += [0] * (sch_len - existing_length)
                self.data.item_excess_amount += [0] * (sch_len - existing_length)

            # fill in from Prev Bill
            if self.prev_bill is not None:
                for item_index, item_qty in enumerate(self.prev_bill.item_qty):
                    if sum(item_qty) > 0:
                        self.item_cmb_ref[item_index].append(-1)  # use -1 as marker for prev abstract
                        self.item_paths[item_index].append([self.data.prev_bill, item_index])
                        self.item_qty[item_index].append(sum(item_qty))  # add total qty from previous bill

            # fill in from measurement items
            for mitem in self.data.mitems:
                item = self.cmbs[mitem[0]][mitem[1]][mitem[2]]
                if not isinstance(item, MeasurementItemHeading):
                    for count, itemno, item_qty in zip(range(len(item.itemnos)), item.itemnos, item.get_total()):
                        item_index = self.schedule.get_item_index(itemno)
                        if item_index is not None:
                            self.item_cmb_ref[item_index].append(mitem[0])
                            self.item_paths[item_index].append(mitem + [count])
                            self.item_qty[item_index].append(item_qty)

            # Evaluate remaining variables from above data
            for item_index in range(sch_len):
                item = self.schedule[item_index]
                # determine total qty
                total_qty = sum(self.item_qty[item_index])
                # determin items above and at normal rates
                if total_qty > item.qty * (1 + 0.01 * item.excess_rate_percent):
                    if item.unit.lower() in INT_ITEMS:
                        self.item_normal_qty[item_index] = math.floor(item.qty * (1 + 0.01 * item.excess_rate_percent))
                    else:
                        self.item_normal_qty[item_index] = round(item.qty * (1 + 0.01 * item.excess_rate_percent), 2)
                    self.item_excess_qty[item_index] = total_qty - self.item_normal_qty[item_index]
                else:
                    self.item_normal_qty[item_index] = total_qty
                # determine amounts
                self.item_normal_amount[item_index] = round(
                    self.item_normal_qty[item_index] * self.data.item_part_percentage[item_index] * 0.01 *
                    self.schedule[item_index].rate, 2)
                self.item_excess_amount[item_index] = round(
                    self.item_excess_qty[item_index] * self.data.item_excess_part_percentage[item_index] * 0.01 *
                    self.data.item_excess_rates[item_index], 2)
                # mbs refered to the bill
                self.cmb_ref = self.cmb_ref | set(self.item_cmb_ref[item_index])  # Add any unique cmb (find union)

            # evaluate total
            self.bill_total_amount = round(sum(self.item_normal_amount) + sum(self.item_excess_amount), 2)
            if self.prev_bill is not None:
                self.bill_since_prev_amount = round(self.bill_total_amount - self.prev_bill.bill_total_amount, 2)
            else:
                self.bill_since_prev_amount = self.bill_total_amount
        elif self.data.bill_type == BILL_CUSTOM:
            # expand variables to schedule length
            existing_length = len(self.data.item_excess_rates)
            if sch_len > existing_length:
                for i in range(sch_len - existing_length):
                    self.data.item_qty.append([0])
                self.item_normal_qty += [0] * (sch_len - existing_length)
                self.item_excess_qty += [0] * (sch_len - existing_length)
                self.data.item_normal_amount += [0] * (sch_len - existing_length)
                self.data.item_excess_amount += [0] * (sch_len - existing_length)
            self.item_qty = self.data.item_qty
            self.item_normal_amount = self.data.item_normal_amount
            self.item_excess_amount = self.data.item_excess_amount
            self.bill_total_amount = round(sum(self.data.item_normal_amount) + sum(self.data.item_excess_amount), 2)
            self.bill_since_prev_amount = self.bill_total_amount

    def get_latex_buffer(self, thisbillpath):
        file_preamble = open(abs_path('latex','preamble.tex'), 'r')
        file_abstractopening = open(abs_path('latex','abstractopening.tex'), 'r')
        file_abstractitem = open(abs_path('latex','abstractitem.tex'), 'r')
        file_abstractitemelement = open(abs_path('latex','abstractitemelement.tex'), 'r')
        file_end = open(abs_path('latex','endabstract.tex'), 'r')

        latex_buffer = file_preamble.read()
        latex_buffer += file_abstractopening.read()  # red the abstract opening

        bill_local_vars = {}  # bill substitution dictionary
        bill_local_vars['$cmbbookno$'] = self.data.cmb_name
        bill_local_vars['$cmbheading$'] = self.data.title
        bill_local_vars['$cmbtitle$'] = 'ABSTRACT OF COST'
        bill_local_vars['$cmbabstractdate$'] = self.data.bill_date
        bill_local_vars['$cmbstartingpage$'] = str(self.data.starting_page)
        bill_local_vars['$cmbbilltotalamount$'] = str(self.bill_total_amount)
        if self.prev_bill is not None:
            bill_local_vars['$cmbbillprevamount$'] = str(self.prev_bill.bill_total_amount)
        else:
            bill_local_vars['$cmbbillprevamount$'] = '0'
        bill_local_vars['$cmbbillsinceprevamount$'] = str(self.bill_since_prev_amount)

        sch_len = self.schedule.length()
        for count, qty_items, cmb_refs, item_paths in zip(range(sch_len), self.item_qty, self.item_cmb_ref,
                                                          self.item_paths):
            if qty_items:  # for every non empty item, include in bill
                # setup required values
                item_local_vars = {}
                item_local_vars_vanilla = {}
                item = self.schedule.get_item_by_index(count)
                if self.item_excess_qty[count] > 0:
                    excess_flag = '\iftrue'
                else:
                    excess_flag = '\iffalse'
                latex_records = ''
                for qty_item, cmb_ref, item_path in zip(qty_items, cmb_refs, item_paths):
                    if qty_item != 0:
                        # setup variables
                        item_record_vars = {}
                        item_record_vars['$cmbqtybf$'] = str(qty_item)
                        item_record_vars['$cmbunit$'] = str(item.unit)
                        if cmb_ref != -1:  # if not prev abstract
                            path_str = str([item_path[0], item_path[1], item_path[2]]) + ':' + str(item_path[3] + 1)
                            item_record_vars['$cmbbf$'] = 'ref:meas:' + path_str
                            item_record_vars['$cmblabel$'] = 'ref:abs:' + path_str
                            item_record_vars['$cmbnormalbillflag$'] = 'iftrue'
                        else:  # if prev abstract
                            if self.prev_bill.data.bill_type == BILL_NORMAL:
                                item_record_vars['$cmbbf$'] = 'ref:abs:abs:' + str([self.data.prev_bill, count])
                                item_record_vars['$cmblabel$'] = ''
                                item_record_vars['$cmbnormalbillflag$'] = 'iftrue'
                            elif self.prev_bill.data.bill_type == BILL_CUSTOM:
                                item_record_vars['$cmbbf$'] = self.prev_bill.data.cmb_name
                                item_record_vars['$cmblabel$'] = ''
                                item_record_vars['$cmbnormalbillflag$'] = 'iffalse'

                        # read in record entries to buffer
                        file_abstractitemelement.seek(0)  # seek file to start
                        latex_record = file_abstractitemelement.read()
                        # make item substitutions in buffer
                        latex_record = replace_all(latex_record, item_record_vars)
                        # add to master buffer
                        latex_records += latex_record

                # add all values to substitution dict
                item_local_vars['$cmbitemno$'] = str(item.itemno)
                item_local_vars['$cmbdescription$'] = str(item.extended_description_limited)
                item_local_vars['$cmbunit$'] = str(item.unit)
                item_local_vars['$cmbrate$'] = str(item.rate)
                item_local_vars['$cmbexcesspercent$'] = str(item.excess_rate_percent)
                item_local_vars['$cmbtotalqty$'] = str(sum(qty_items))
                item_local_vars['$cmbnormalqty$'] = str(self.item_normal_qty[count])
                item_local_vars['$cmbexcessqty$'] = str(self.item_excess_qty[count])
                item_local_vars['$cmbexcessrate$'] = str(self.data.item_excess_rates[count])
                item_local_vars['$cmbnormalpr$'] = str(
                    round(self.data.item_part_percentage[count] * 0.01 * self.schedule[count].rate, 2))
                item_local_vars['$cmbexcesspr$'] = str(
                    round(self.data.item_excess_part_percentage[count] * 0.01 * self.data.item_excess_rates[count], 2))
                item_local_vars['$cmbnormalamount$'] = str(self.item_normal_amount[count])
                item_local_vars['$cmbexcessamount$'] = str(self.item_excess_amount[count])
                item_local_vars['$cmbabslabel$'] = 'ref:abs:abs:' + str(thisbillpath + [count])

                item_local_vars_vanilla['$cmbexcessflag$'] = excess_flag
                item_local_vars_vanilla['$cmbrecords$'] = latex_records

                # write entries
                file_abstractitem.seek(0)  # seek to start
                latex_buffer += file_abstractitem.read()  # read item template
                # make item substitutions
                latex_buffer = replace_all(latex_buffer, item_local_vars)
                latex_buffer = replace_all_vanilla(latex_buffer, item_local_vars_vanilla)

        latex_buffer += file_end.read()
        latex_buffer = replace_all(latex_buffer, bill_local_vars)

        # close all files
        file_preamble.close()
        file_abstractopening.close()
        file_abstractitem.close()
        file_abstractitemelement.close()
        file_end.close()

        return latex_buffer

    def get_latex_buffer_bill(self):
        file_billopening = open(abs_path('latex','billopening.tex'), 'r')
        file_billitem = open(abs_path('latex','billitem.tex'), 'r')
        file_end = open(abs_path('latex','endbill.tex'), 'r')

        latex_buffer = file_billopening.read()

        bill_local_vars = {}  # bill substitution dictionary
        bill_local_vars['$cmbheading$'] = self.data.title
        bill_local_vars['$cmbabstractdate$'] = self.data.bill_date
        bill_local_vars['$cmbbilltotalamount$'] = str(self.bill_total_amount)
        if self.prev_bill is not None:
            bill_local_vars['$cmbbillprevamount$'] = str(self.prev_bill.bill_total_amount)
        else:
            bill_local_vars['$cmbbillprevamount$'] = '0'
        bill_local_vars['$cmbbillsinceprevamount$'] = str(self.bill_since_prev_amount)

        sch_len = self.schedule.length()
        for count, qty_items, item_paths in zip(range(sch_len), self.item_qty,
                                                          self.item_paths):
            if self.prev_bill is not None:
                sprev_item_normal_amount = self.item_normal_amount[count] - \
                                           self.prev_bill.item_normal_amount[count]
                sprev_item_excess_amount = self.item_excess_amount[count] - \
                                           self.prev_bill.item_excess_amount[count]
            else:
                sprev_item_normal_amount = self.item_normal_amount[count]
                sprev_item_excess_amount = self.item_excess_amount[count]

            if qty_items:  # for every non empty item, include in bill
                # setup required values
                item_local_vars = {}
                item_local_vars_vanilla = {}
                item = self.schedule.get_item_by_index(count)
                if self.item_excess_qty[count] > 0:
                    excess_flag = '\iftrue'
                else:
                    excess_flag = '\iffalse'

                # add all values to substitution dict
                item_local_vars['$cmbitemno$'] = str(item.itemno)
                item_local_vars['$cmbdescription$'] = str(item.extended_description_limited)
                item_local_vars['$cmbunit$'] = str(item.unit)
                item_local_vars['$cmbrate$'] = str(item.rate)
                item_local_vars['$cmbexcesspercent$'] = str(item.excess_rate_percent)
                item_local_vars['$cmbtotalqty$'] = str(sum(qty_items))
                item_local_vars['$cmbnormalqty$'] = str(self.item_normal_qty[count])
                item_local_vars['$cmbexcessqty$'] = str(self.item_excess_qty[count])
                item_local_vars['$cmbexcessrate$'] = str(self.data.item_excess_rates[count])
                item_local_vars['$cmbnormalpr$'] = str(
                    round(self.data.item_part_percentage[count] * 0.01 * self.schedule[count].rate, 2))
                item_local_vars['$cmbexcesspr$'] = str(
                    round(self.data.item_excess_part_percentage[count] * 0.01 * self.data.item_excess_rates[count], 2))
                item_local_vars['$cmbnormalamount$'] = str(self.item_normal_amount[count])
                item_local_vars['$cmbexcessamount$'] = str(self.item_excess_amount[count])
                item_local_vars['$cmbnormalsinceprevamount$'] = str(sprev_item_normal_amount)
                item_local_vars['$cmbexcesssinceprevamount$'] = str(sprev_item_excess_amount)

                item_local_vars_vanilla['$cmbexcessflag$'] = excess_flag

                # write entries
                file_billitem.seek(0)  # seek to start
                latex_buffer += file_billitem.read()  # read item template
                # make item substitutions
                latex_buffer = replace_all(latex_buffer, item_local_vars)
                latex_buffer = replace_all_vanilla(latex_buffer, item_local_vars_vanilla)

        latex_buffer += file_end.read()
        latex_buffer = replace_all(latex_buffer, bill_local_vars)

        # close all files
        file_billopening.close()
        file_billitem.close()
        file_end.close()

        return latex_buffer

    def export_ods_bill(self,filename,project_settings_dict):
        spreadsheet = Workbook()

        # Sheet 1
        sheet = spreadsheet.active
        sheet.title = 'Data'
        sheet_head = ['Agmnt.No','Description','Unit','Total Qty','Below Dev Qty',
                      'Above Dev Qty','Agmnt FR','Agmnt PR','Excess FR','Excess FR',
                      'Total Bel Dev','Total Above Dev',
                      'Since Prev below Dev','Since Prev Above Dev','Dev Limit']
        for c,head in enumerate(sheet_head):
            sheet.cell(row=1,column=c+1).value = head
            sheet.cell(row=1,column=c+1).font = Font(bold=True)
            sheet.cell(row=1,column=c+1).alignment = Alignment(wrap_text=True, vertical='center')
        for count in range(self.schedule.length()):
            if self.prev_bill is not None:
                sprev_item_normal_amount = self.item_normal_amount[count] - \
                                           self.prev_bill.item_normal_amount[count]
                sprev_item_excess_amount = self.item_excess_amount[count] - \
                                           self.prev_bill.item_excess_amount[count]
            else:
                sprev_item_normal_amount = self.item_normal_amount[count]
                sprev_item_excess_amount = self.item_excess_amount[count]

            item = self.schedule.get_item_by_index(count)
            sheet.cell(row=count+2, column=1).value = item.itemno
            sheet.cell(row=count+2, column=2).value = item.description
            sheet.cell(row=count+2, column=2).alignment = Alignment(wrap_text=True)
            sheet.cell(row=count+2, column=3).value = item.unit
            sheet.cell(row=count+2, column=4).value = sum(self.item_qty[count])
            sheet.cell(row=count+2, column=5).value = self.item_normal_qty[count]
            sheet.cell(row=count+2, column=6).value = self.item_excess_qty[count]
            sheet.cell(row=count+2, column=7).value = item.rate
            sheet.cell(row=count+2, column=8).value = round(self.data.item_part_percentage[count] *
                                             0.01 * self.schedule[count].rate, 2)
            sheet.cell(row=count+2, column=9).value = self.data.item_excess_rates[count]
            sheet.cell(row=count+2, column=10).value = round(self.data.item_excess_part_percentage[count] *
                                             0.01 * self.data.item_excess_rates[count], 2)
            sheet.cell(row=count+2, column=11).value = self.item_normal_amount[count]
            sheet.cell(row=count+2, column=12).value = self.item_excess_amount[count]
            sheet.cell(row=count+2, column=13).value = sprev_item_normal_amount
            sheet.cell(row=count+2, column=14).value = sprev_item_excess_amount
            sheet.cell(row=count+2, column=15).value = item.excess_rate_percent

        # sheet formatings
        sheet.column_dimensions['B'].width = 50
        sheet.page_setup.orientation = worksheet.Worksheet.ORIENTATION_LANDSCAPE
        sheet.page_setup.paperSize = worksheet.Worksheet.PAPERSIZE_LEGAL
        sheet.page_setup.fitToHeight = 99
        sheet.page_setup.fitToWidth = 1

        # Sheet 2 Deviation statement
        sheet2 = spreadsheet.create_sheet()
        sheet2.title = 'Deviation'
        template = load_workbook(filename = abs_path('ods_templates','dev.xlsx'))
        rowend = 12
        colend = 18
        rowend_end = 3
        # Copy all from dev start
        template_start_sheet = template.get_sheet_by_name('start')
        for row in range(1,rowend+1):
            for column in range(1,colend+1):
                sheet2.cell(row=row, column=column).value = template_start_sheet.cell(row=row, column=column).value
                sheet2.cell(row=row, column=column).style = template_start_sheet.cell(row=row, column=column).style
        # copy all values
        for count in range(self.schedule.length()):
            item = self.schedule.get_item_by_index(count)
            sheet2.cell(row=count+rowend+1, column=1).value = '=IF(INDIRECT(ADDRESS(ROW(),5))<>0,MAX(INDIRECT("A$12:" & ADDRESS(ROW()-1,1)))+1,"")'
            sheet2.cell(row=count+rowend+1, column=2).value = item.itemno
            sheet2.cell(row=count+rowend+1, column=3).value = item.description
            if item.qty > 0:
                percent_dev = round((sum(self.item_qty[count]) - item.qty)/item.qty*100,2) if item.qty != 0 else 0
                # Fill in values
                sheet2.cell(row=count+rowend+1, column=4).value = item.unit
                sheet2.cell(row=count+rowend+1, column=5).value = item.qty
                sheet2.cell(row=count+rowend+1, column=6).value = sum(self.item_qty[count])
                sheet2.cell(row=count+rowend+1, column=7).value = sum(self.item_qty[count]) - item.qty
                sheet2.cell(row=count+rowend+1, column=8).value = percent_dev
                if self.item_normal_qty[count]-item.qty > 0:
                    sheet2.cell(row=count+rowend+1, column=9).value = self.item_normal_qty[count] - item.qty
                if self.item_excess_qty[count] > 0:
                    sheet2.cell(row=count+rowend+1, column=10).value = self.item_excess_qty[count]
                sheet2.cell(row=count+rowend+1, column=11).value = \
                        '=IF(ABS(INDIRECT(ADDRESS(ROW(),8)))>' + str(DEV_LIMIT_STATEMENT) + ',' + \
                        str(self.item_normal_qty[count] - item.qty) + ',0)'
                sheet2.cell(row=count+rowend+1, column=12).value = item.rate
                sheet2.cell(row=count+rowend+1, column=15).value = self.data.item_excess_rates[count]
                # Fill in formulas
                sheet2.cell(row=count+rowend+1, column=13).value = \
                    '=INDIRECT(ADDRESS(ROW(),11))*INDIRECT(ADDRESS(ROW(),12))'
                sheet2.cell(row=count+rowend+1, column=14).value = '=INDIRECT(ADDRESS(ROW(),10))'
                sheet2.cell(row=count+rowend+1, column=16).value = \
                    '=INDIRECT(ADDRESS(ROW(),14))*INDIRECT(ADDRESS(ROW(),15))'
                sheet2.cell(row=count+rowend+1, column=17).value = \
                    '=ABS(INDIRECT(ADDRESS(ROW(),13)))+ABS(INDIRECT(ADDRESS(ROW(),16)))'
            # formatings
            sheet2.cell(row=count+rowend+1, column=1).alignment = Alignment(horizontal='center')
            sheet2.cell(row=count+rowend+1, column=2).alignment = Alignment(horizontal='center')
            sheet2.cell(row=count+rowend+1, column=3).alignment = Alignment(wrap_text=True)
         # Copy all from dev end
        template_end_sheet = template.get_sheet_by_name('end')
        for row,row_ in zip(range(rowend+self.schedule.length()+1,rowend+self.schedule.length()+rowend_end+1),range(1,rowend_end+1)):
            for column,column_ in zip(range(1,colend+1),range(1,colend+1)):
                sheet2.cell(row=row, column=column).value = template_end_sheet.cell(row=row_, column=column_).value
                sheet2.cell(row=row, column=column).style = template_end_sheet.cell(row=row_, column=column_).style
        sheet2.cell(row=rowend+self.schedule.length()+1, column=colend-1).value = \
            '=SUM('+get_column_letter(colend-1)+str(rowend+1)+':'+get_column_letter(colend-1)+\
            str(rowend+self.schedule.length())+')'
        # sheet2 formatings
        for column in range(1,colend+1):
            # copy coumn widths
            sheet2.column_dimensions[get_column_letter(column)].width = \
                template_start_sheet.column_dimensions[get_column_letter(column)].width
        sheet2.page_setup.orientation = worksheet.Worksheet.ORIENTATION_LANDSCAPE
        sheet2.page_setup.paperSize = worksheet.Worksheet.PAPERSIZE_LEGAL
        sheet2.page_setup.fitToHeight = 99
        sheet2.page_setup.fitToWidth = 1
        sheet2.print_options.horizontalCentered = True

        # Copy all data
        sheet2['C3'] = project_settings_dict["$cmbnameofwork$"]
        sheet2['C4'] = project_settings_dict["$cmbagency$"]
        sheet2['C5'] = project_settings_dict["$cmbagmntno$"]

        # Save Document
        spreadsheet.save(filename)

    def get_text(self):
        total = [self.bill_total_amount, self.bill_since_prev_amount]
        if self.data.bill_type == BILL_NORMAL:
            return '<b>' + clean_markup(self.data.title) + '</b> | CMB.No.<b>' + clean_markup(
                self.data.cmb_name) + ' dated ' + clean_markup(self.data.bill_date) + '</b> | TOTAL: <b>' + str(
                total) + '</b>'
        if self.data.bill_type == BILL_CUSTOM:
            return '<span foreground="red"><b>' + clean_markup(self.data.title) + '</b> | CMB.No.<b>' + clean_markup(
                self.data.cmb_name) + ' dated ' + clean_markup(self.data.bill_date) + '</b> | TOTAL: <b>' + str(
                total) + '</b></span>'


    def print_item(self):
        print("bill " + self.data.title + " start")
        for count in range(self.schedule.length()):
            itemno = self.schedule[count].itemno
            desc = self.schedule[count].description
            unit = self.schedule[count].unit
            if self.data.bill_type == BILL_NORMAL:
                rate = [self.schedule[count].rate, self.data.item_excess_rates[count]]
            elif self.data.bill_type == BILL_CUSTOM:
                rate = [self.schedule[count].rate]
            amount = [self.item_normal_amount[count], self.item_excess_amount[count]]
            print([itemno, desc, unit, rate, amount])
        print(['\nTotal Amount | Since Prev ', [self.bill_total_amount, self.bill_since_prev_amount]])
        print("bill end")


class BillView:
    # Bill view callbacks

    # for unselecting on escape
    def onKeyPressTreeviewSchedule(self, treeview, event):
        keyname = Gdk.keyval_name(event.keyval)
        if keyname == "Escape":
            self.tree.get_selection().unselect_all()

    # Bill view methods

    def add_bill(self):
        toplevel = self.tree.get_toplevel()  # get toplevel window
        bill_dialog = BillDialog(toplevel, BillData(), self)
        data = bill_dialog.run()
        if data is not None:  # if cancel not pressed
            bill = Bill(self)
            bill.set_modal(data)
            self.insert_item_at_row(bill, None)

    def add_bill_custom(self):
        toplevel = self.tree.get_toplevel()  # get toplevel window
        bill_dialog = BillDialog(toplevel, BillData(BILL_CUSTOM), self)
        data = bill_dialog.run()
        if data is not None:  # if cancel not pressed
            bill = Bill(self)
            bill.set_modal(data)
            self.insert_item_at_row(bill, None)

    def insert_item_at_selection(self, item):
        selection = self.tree.get_selection()
        if selection.count_selected_rows() != 0:  # if selection exists copy at selection
            [model, paths] = selection.get_selected_rows()
            path = paths[0].get_indices()
            row = path[0]
            self.insert_item_at_row(item, row)
        else:  # if no selection append a new item
            self.insert_item_at_row(item, None)

    @undoable
    def insert_item_at_row(self, item, row):  # note needs rows to be sorted
        if row is not None:
            new_row = row
            self.bills.insert(row, item)
        else:
            self.bills.append(item)
            new_row = len(self.bills) - 1
        self.update_store()

        yield "Insert data items to bill at row '{}'".format(new_row)
        # Undo action
        self.delete_row(new_row)
        self.update_store()

    def edit_selected_row(self):
        selection = self.tree.get_selection()
        if selection.count_selected_rows() != 0:  # if selection exists copy at selection
            [model, paths] = selection.get_selected_rows()
            path = paths[0].get_indices()
            row = path[0]
            bill = self.bills[row]
            data = copy.deepcopy(bill.get_modal())  # make a copy of the modal

            toplevel = self.tree.get_toplevel()  # get toplevel window
            bill_dialog = BillDialog(toplevel, data, self, row)
            new_data = bill_dialog.run()
            if new_data is not None:  # if cancel not pressed
                self.edit_item_at_row(new_data, row)

    @undoable
    def edit_item_at_row(self, data, row):
        if row is not None:
            old_data = copy.deepcopy(self.bills[row].get_modal())
            self.bills[row].set_modal(data)
        self.update_store()

        yield "Edit bill item at row '{}'".format(row)
        # Undo action
        if row is not None:
            self.bills[row].set_modal(old_data)
        self.update_store()


    def delete_selected_row(self):
        # get rows to delete
        selection = self.tree.get_selection()
        [model, path] = selection.get_selected_rows()
        row = path[0].get_indices()[0]
        # delete row
        self.delete_row(row)

    @undoable
    def delete_row(self, row):
        item = self.bills[row]
        del self.bills[row]
        self.update_store()

        yield "Delete data items from bill at row '{}'".format(row)
        # Undo action
        self.insert_item_at_row(item, row)
        self.update_store()

    def copy_selection(self):
        selection = self.tree.get_selection()
        if selection.count_selected_rows() != 0:  # if selection exists
            [model, paths] = selection.get_selected_rows()
            path = paths[0]
            row = int(path.get_indices()[0])  # get item row
            item = self.bills[row]
            data = item.get_modal()
            text = pickle.dumps(data)  # dump item as text
            self.clipboard.set_text(text, -1)  # push to clipboard
        else:  # if no selection
            print("No items selected to copy")

    def paste_at_selection(self):
        text = self.clipboard.wait_for_text()  # get text from clipboard
        if text is not None:
            try:
                data = pickle.loads(text)  # recover item from string
                if isinstance(data, BillData):
                    selection = self.tree.get_selection()
                    if selection.count_selected_rows() != 0:  # if selection exists copy at selection
                        [model, paths] = selection.get_selected_rows()
                        path = paths[0].get_indices()
                        row = path[0]
                        
                        # Handle different bill types
                        if self.bills[row].data.bill_type == BILL_CUSTOM:
                            if data.bill_type == BILL_NORMAL:
                                # create duplicate bill
                                bill = Bill(self)
                                bill.set_modal(data)
                                bill.update_values()
                                # Fill in values for custom bill from duplicate bill
                                data.item_normal_amount = bill.item_normal_amount
                                data.item_excess_amount = bill.item_excess_amount
                                data.item_qty = []
                                for qtys in bill.item_qty:
                                    data.item_qty.append([sum(qtys)])
                                # clear items for normal bill
                                data.mitems = []
                                data.item_part_percentage = []  # part rate for exess rate items
                                data.item_excess_part_percentage = []  # part rate for exess rate items
                                data.item_excess_rates = []  # list of excess rates above excess_percentage
                                # set bill type
                                data.bill_type = BILL_CUSTOM
                        else:
                            data.mitems = []  # clear measured items
                            # clear additional elements for custom bill
                            data.item_qty = []  # qtys of items b/f
                            data.item_normal_amount = []  # total item amount for qty at normal rate
                            data.item_excess_amount = []  # amounts for qty at excess rate
                            # set bill type
                            data.bill_type = BILL_NORMAL
                            
                        self.edit_item_at_row(data, row)
                    else:  # if selection do not exist paste at end
                        data.mitems = []  # clear measured items
                        bill = Bill(self)
                        bill.set_modal(data)
                        self.insert_item_at_row(bill, None)
            except:
                print('No valid data in clipboard')
        else:
            print("No text on the clipboard.")

    def undo(self):
        setstack(self.stack)  # select bill undo stack
        print self.stack.undotext()
        self.stack.undo()

    def redo(self):
        setstack(self.stack)  # select bill undo stack
        print self.stack.redotext()
        self.stack.redo()

    def get_data_object(self):
        data = []
        for bill in self.bills:
            data.append(bill.get_modal())
        return data

    def set_data_object(self, schedule, measurement_view, datalist):
        self.bills = []
        self.schedule = schedule
        self.measurement_view = measurement_view
        self.cmbs = measurement_view.cmbs
        for data in datalist:
            bill = Bill(self)
            bill.set_modal(data)
            self.bills.append(bill)
        self.update_store()

    def clear(self):
        self.stack.clear()
        self.bills = []
        self.update_store()

    def update_store(self):

        # Update all bills
        for bill in self.bills:
            bill.update_values()

        # Update measurement flags
        ManageResourses().update_billed_flags()

        # Update Bill Store
        self.store.clear()
        for count, bill in enumerate(self.bills):
            self.store.append()
            self.store[count][0] = str(count + 1)
            self.store[count][1] = bill.get_text()
        setstack(self.stack)  # select undo stack

    def render_selected(self, folder, replacement_dict):
        # get selection
        selection = self.tree.get_selection()
        if selection.count_selected_rows() != 0 and folder is not None:  # if selection exists
            # get path of selection
            [model, paths] = selection.get_selected_rows()
            path = paths[0].get_indices()
            code = self.render(folder, replacement_dict, path)
            return code
        else:
            return [CMB_WARNING, 'Please select a Bill for rendering']

    def render(self, folder, replacement_dict, path, recursive=True):
        if self.bills[path[0]].data.bill_type == BILL_NORMAL:  # render only if normal bill
            bill = self.bills[path[0]]
            # fill in latex buffer
            bill.update_values()  # build all data structures
            if bill.prev_bill is not None:
                bill.prev_bill.update_values()
            latex_buffer = bill.get_latex_buffer([path[0]])
            latex_buffer_bill = bill.get_latex_buffer_bill()
            # make global variables replacements
            latex_buffer = replace_all(latex_buffer, replacement_dict)
            latex_buffer_bill = replace_all(latex_buffer_bill, replacement_dict)

            # include linked cmbs
            replacement_dict_cmbs = {}
            external_docs = ''
            for cmbpath in bill.cmb_ref:
                if cmbpath != -1:
                    external_docs += '\externaldocument{cmb_' + str(cmbpath + 1) + '}\n'
                elif bill.data.prev_bill is not None: # prev abstract
                    external_docs += '\externaldocument{abs_' + str(bill.data.prev_bill + 1) + '}\n'
            replacement_dict_cmbs['$cmbexternaldocs$'] = external_docs
            latex_buffer = replace_all_vanilla(latex_buffer, replacement_dict_cmbs)

            # write output
            filename = posix_path(folder, 'abs_' + str(path[0] + 1) + '.tex')
            file_latex = open(filename, 'w')
            file_latex.write(latex_buffer)
            file_latex.close()

            filename_bill = posix_path(folder, 'bill_' + str(path[0] + 1) + '.tex')
            file_latex_bill = open(filename_bill, 'w')
            file_latex_bill.write(latex_buffer_bill)
            file_latex_bill.close()

            filename_bill_ods = posix_path(folder, 'bill_' + str(path[0] + 1) + '.xlsx')

            # run latex on file and dependencies
            if recursive:  # if recursive call
                # Render all cmbs depending on the bill
                for cmb_ref in bill.cmb_ref:
                    if cmb_ref is not -1:  # if not prev bill
                        code = self.measurement_view.render(folder, replacement_dict, self.bills, [cmb_ref], False)
                        if code[0] == CMB_ERROR:
                            return code
                # Render prev bill
                if bill.prev_bill is not None and bill.prev_bill.data.bill_type == BILL_NORMAL:
                    code = self.render(folder, replacement_dict, [bill.data.prev_bill], False)
                    if code[0] == CMB_ERROR:
                        return code

            # Render this bill
            code = run_latex(posix_path(folder), filename)
            if code == CMB_ERROR:
                return [CMB_ERROR, 'Rendering of Bill: ' + self.bill.data.title + ' failed']
            code_bill = run_latex(posix_path(folder), filename_bill)
            if code_bill == CMB_ERROR:
                return [CMB_ERROR, 'Rendering of Bill Schedule: ' + self.bill.data.title + ' failed']

            # Render all cmbs again to rebuild indexes on recursive run
            if recursive:  # if recursive call
                for cmb_ref in bill.cmb_ref:
                    if cmb_ref is not -1:  # if not prev bill
                        code = self.measurement_view.render(folder, replacement_dict, self.bills, [cmb_ref], False)
                        if code[0] == CMB_ERROR:
                            return code
                bill.export_ods_bill(filename_bill_ods,replacement_dict)

            return [CMB_INFO, 'Bill: ' + self.bills[path[0]].data.title + ' rendered successfully']
        else:
            return [CMB_WARNING, 'Rendering of custom bill not supported']

    def __init__(self, schedule, measurement_view, list_store_object, tree_view_object):
        self.schedule = schedule
        self.measurement_view = measurement_view
        self.store = list_store_object
        self.tree = tree_view_object
        self.cmbs = measurement_view.cmbs

        self.bills = []
        self.stack = Stack()  # initialise undo/redo stack
        setstack(self.stack)  # select bill undo stack

        self.clipboard = Gtk.Clipboard.get(Gdk.SELECTION_CLIPBOARD)  # initialise clipboard

        self.tree.connect("key-press-event", self.onKeyPressTreeviewSchedule)
