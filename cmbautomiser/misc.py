#!/usr/bin/env python
# -*- coding: utf-8 -*-
#
# misc.py
#  
#  Copyright 2014 Manu Varkey <manuvarkey@gmail.com>
#  
#  This program is free software; you can redistribute it and/or modify
#  it under the terms of the GNU General Public License as published by
#  the Free Software Foundation; either version 2 of the License, or
#  (at your option) any later version.
#  
#  This program is distributed in the hope that it will be useful,
#  but WITHOUT ANY WARRANTY; without even the implied warranty of
#  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#  GNU General Public License for more details.
#  
#  You should have received a copy of the GNU General Public License
#  along with this program; if not, write to the Free Software
#  Foundation, Inc., 51 Franklin Street, Fifth Floor, Boston,
#  MA 02110-1301, USA.
#  
#  

import subprocess, threading, os, posixpath, platform, logging

from gi.repository import Gtk, Gdk, GLib, Pango
from urllib.parse import urlparse
from urllib.request import url2pathname
import openpyxl

# Setup logger object
log = logging.getLogger(__name__)

## GLOBAL CONSTANTS

# Program name
PROGRAM_NAME = 'CMB Automiser'
# Item codes for data types
MEAS_NO = 1
MEAS_L = 2
MEAS_DESC = 3
MEAS_CUST = 4
# CMB error codes used for displaying info in main window
CMB_ERROR = -1
CMB_WARNING = -2
CMB_OK = 0
CMB_INFO = 0
# For tracking window management
CHILD_WINDOW = 1
PARENT_WINDOW = 0
main_hidden = 0
child_windows = 0
# Used in bill module for indicating type of bill
BILL_CUSTOM = 1
BILL_NORMAL = 2
# background colors for treeview
MEAS_COLOR_LOCKED = '#BABDB6'
MEAS_COLOR_NORMAL = '#FFFFFF'
MEAS_COLOR_SELECTED = '#729FCF'
# Timeout for killing Latex subprocess
LATEX_TIMEOUT = 300 # 5 minutes
# Item description wrap-width for screen purpose
CMB_DESCRIPTION_WIDTH = 60
CMB_DESCRIPTION_MAX_LENGTH = 1000
# Deviation statement
DEV_LIMIT_STATEMENT = 10
# List of units which will be considered as integer values
INT_ITEMS = ['point', 'points', 'pnt', 'pnts', 'number', 'numbers', 'no', 'nos', 'lot', 'lots',
             'lump', 'lumpsum', 'lump-sum', 'lump sum', 'ls', 'each','job','jobs','set','sets',
             'pair','pairs',
             'pnt.', 'no.', 'nos.', 'l.s.', 'l.s']
# String used for checking file version
PROJECT_FILE_VER = 'CMBAUTOMISER_FILE_REFERENCE_VER_3'
# Item codes for project global variables
global_vars = ['$cmbnameofwork$',
               '$cmbagency$',
               '$cmbagmntno$', 
               '$cmbsituation$',
               '$cmbdateofstart$',
               '$cmbdateofstartasperagmnt$',
               '$cmbissuedto$',
               '$cmbvarifyingauthority$',
               '$cmbvarifyingauthorityoffice$',
               '$cmbissuingauthority$',
               '$cmbissuingauthorityoffice$']
global_vars_captions = ['Name of Work', 
                        'Agency',
                        'Agreement Number',
                        'Situation',
                        'Date of Start',
                        'Date of start as per Agmnt.',
                        'CMB Issued to',
                        'Varifying Authority',
                        'Varifying Authority Office',
                        'Issuing Authority',
                        'Issuing Authority Office']
               
## GLOBAL VARIABLES

# Dict for storing saved settings
global_settings_dict = dict()

def set_global_platform_vars():
    """Setup global platform dependent variables"""
    
    if platform.system() == 'Linux':
        global_settings_dict['latex_path'] = 'pdflatex'
    elif platform.system() == 'Windows':
        global_settings_dict['latex_path'] = abs_path(
                    'miketex\\miktex\\bin\\pdflatex.exe')

## GLOBAL CLASSES

class UserEntryDialog():
    """Creates a dialog box for entry of custom data fields
    
        Arguments:
            parent: Parent Window
            window_caption: Window Caption to be displayed on Dialog
            item_values: Item values to be requested from user
            item_captions: Description of item values to be shown to user
    """
    
    def __init__(self, parent, window_caption, item_values, item_captions):
        self.toplevel = parent
        self.entrys = []
        self.item_values = item_values
        self.item_captions = item_captions

        self.dialog_window = Gtk.Dialog(window_caption, parent, Gtk.DialogFlags.MODAL,
            (Gtk.STOCK_CANCEL, Gtk.ResponseType.CANCEL,
             Gtk.STOCK_OK, Gtk.ResponseType.OK))
        self.dialog_window.set_title(window_caption)
        self.dialog_window.set_resizable(True)
        self.dialog_window.set_border_width(5)
        self.dialog_window.set_size_request(int(self.toplevel.get_size_request()[0]*0.8),-1)
        self.dialog_window.set_default_response(Gtk.ResponseType.OK)

        # Pack Dialog
        dialog_box = self.dialog_window.get_content_area()
        grid = Gtk.Grid()
        grid.set_column_spacing(5)
        grid.set_row_spacing(5)
        grid.set_border_width(5)
        grid.set_hexpand(True)
        dialog_box.add(grid)
        for caption in self.item_captions:
            # Captions
            user_label = Gtk.Label(caption)
            user_label.set_halign(Gtk.Align.END)
            # Text Entry
            user_entry = Gtk.Entry()
            user_entry.set_hexpand(True)
            user_entry.set_activates_default(True)
            # Pack Widgets
            grid.attach_next_to(user_label, None, Gtk.PositionType.BOTTOM, 1, 1)
            grid.attach_next_to(user_entry, user_label, Gtk.PositionType.RIGHT, 1, 1)
            self.entrys.append(user_entry)
        # Add data
        for value, user_entry in zip(self.item_values, self.entrys):
            user_entry.set_text(value)
                
    def run(self):
        """Display dialog box and modify Item Values in place
        
            Save modified values to "item_values" (item passed by reference)
            if responce is Ok. Discard modified values if response is Cancel.
            
            Returns:
                True on Ok
                False on Cancel
        """
        # Run dialog
        self.dialog_window.show_all()
        response = self.dialog_window.run()
        
        if response == Gtk.ResponseType.OK:
            # Get formated text and update item_values
            for key, user_entry in zip(range(len(self.item_values)), self.entrys):
                cell = user_entry.get_text()
                try:  # try evaluating string
                    if type(self.item_values[key]) is str:
                        cell_formated = str(cell)
                    elif type(self.item_values[key]) is int:
                        cell_formated = str(float(cell))
                    elif type(self.item_values[key]) is float:
                        cell_formated = str(int(cell))
                    else:
                        cell_formated = ''
                except:
                    cell_formated = ''
                self.item_values[key] = cell_formated
                
            self.dialog_window.destroy()
            return True
        else:
            self.dialog_window.destroy()
            return False
            
class SpreadsheetDialog:
    """Dialog for manage input and output of spreadsheets"""
   
    def __init__(self, parent, filename, columntypes, captions, dimensions = None):
        """Initialise SpreadsheetDialog class
        
            Arguments:
                parent: Parent widget (Main window)
                filename: 
                columntypes: Data types of columns. 
                             Takes following values:
                                misc.MEAS_NO: Integer
                                misc.MEAS_L: Float
                                misc.MEAS_DESC: String
                                misc.MEAS_CUST: Value omited
                dimensions: List of two lists passing column widths and expand properties
        """
        log.info('SpreadsheetDialog - Initialise')
        # Setup variables
        self.parent = parent
        self.filename = filename
        self.captions = captions
        self.columntypes = columntypes
        self.dimensions = dimensions
        
        self.top = 0
        self.bottom = 0
        self.left = 0
        self.right = 0
        self.values = []
        self.spreadsheet = None
        self.sheet = ''

        # Setup dialog window
        self.builder = Gtk.Builder()
        self.builder.add_from_file(abs_path("interface","spreadsheetdialog.glade"))
        self.window = self.builder.get_object("dialog")
        self.window.set_transient_for(self.parent)
        self.window.set_default_size(950,500)
        self.builder.connect_signals(self)

        # Get required objects
        self.combo = self.builder.get_object("combobox_sheet")
        self.combo_store = self.builder.get_object("liststore_combo")
        self.tree = self.builder.get_object("treeview_schedule")
        self.entry_top = self.builder.get_object("entry_top")
        self.entry_bottom = self.builder.get_object("entry_bottom")
        self.entry_left = self.builder.get_object("entry_left")
        self.entry_right = self.builder.get_object("entry_right")
        
        # Setup treeview
        self.columns = []
        self.cells = []
        # Setup row number column
        cell_row = Gtk.CellRendererText()
        column_row = Gtk.TreeViewColumn('', cell_row)
        column_row.add_attribute(cell_row, "markup", 0)
        column_row.set_min_width(50)
        column_row.set_fixed_width(50)
        cell_row.props.wrap_width = 50
        cell_row.props.background = MEAS_COLOR_LOCKED
        self.cells.append(cell_row)
        self.columns.append(column_row)
        self.tree.append_column(column_row)
        # Setup remaining columns
        for c_no, caption  in enumerate(self.captions,1):
            cell = Gtk.CellRendererText()
            column = Gtk.TreeViewColumn(caption, cell)
            column.add_attribute(cell, "text", c_no)
            self.cells.append(cell)
            self.columns.append(column)
            self.tree.append_column(column)
        # Setup dimensions
        if dimensions is not None:
            self.setup_column_props(*dimensions)
        # Setup liststore model
        types = [str] + [str]*len(self.columntypes)
        self.store = Gtk.ListStore(*types)
        self.tree.set_model(self.store)
        # Misc options
        self.tree.set_grid_lines(3)
        self.tree.set_enable_search(True)
        search_cols = [no for no,x in enumerate(self.columntypes,1) if x == MEAS_DESC]
        self.tree.set_search_equal_func(self.equal_func, [0,1,2,3])
        
        # Read file into spreadsheet object
        if filename is not None:
            try:
                self.spreadsheet = Spreadsheet(filename)
            except:
                self.spreadsheet = None
                log.warning('SpreadsheetDialog - Spreadsheet could not be read - ' + filename)
                
        # Setup combobox
        if self.spreadsheet:
            sheets = self.spreadsheet.sheets()
            for sheet in sheets:
                self.combo_store.append([sheet])
            if sheets:
                self.combo.set_active_id(sheets[0])
                self.update()
            

    def run(self):
        """Display dialog box and return data model
        
            Returns:
                Data Model on Ok
                [] on Cancel
        """
        self.window.show_all()
        response = self.window.run()
        self.window.destroy()

        if response == 1 and self.spreadsheet:
            log.info('SpreadsheetDialog - run - Response Ok')
            return self.values
        else:
            log.info('SpreadsheetDialog - run - Response Cancel')
            return []
    
    def update(self):
        """Update contents from input values"""
        log.info('SpreadsheetDialog - Update')
        
        # Read if sheet changed
        sheet = self.combo_store[self.combo.get_active_iter()][0]
        if sheet != self.sheet:
            self.sheet = sheet
            self.spreadsheet.set_active_sheet(self.sheet)
            self.entry_top.set_text('1')
            self.entry_bottom.set_text(str(self.spreadsheet.length()+1))
            self.entry_left.set_text('1')
            
        # Read values of entries
        self.top = int(self.entry_top.get_text())
        self.bottom = int(self.entry_bottom.get_text())
        self.left = int(self.entry_left.get_text())
        
        # Set values
        self.entry_right.set_text(str(self.left + len(self.columntypes)))
        
        # Read spreadsheet
        self.values = self.spreadsheet.read_rows(self.columntypes, start=self.top-1, end=self.bottom-1, left=self.left-1)
        
        # Update store
        self.store.clear()
        for slno, value in enumerate(self.values, self.top):
            self.store.append(['<b>' + str(slno) + '</b>'] + value)
                
    def setup_column_props(self, widths, expandables):
        """Set column properties
            Arguments:
                widths: List of column widths type-> [int, ...]. None values are skiped.
                expandables: List of expand property type-> [bool, ...]. None values are skiped.
        """
        for column, cell, width, expandable in zip(self.columns[1:], self.cells[1:], widths, expandables):
            if width != None:
                column.set_min_width(width)
                column.set_fixed_width(width)
                cell.props.wrap_width = width
            if expandable != None:
                column.set_expand(expandable)
    
    def equal_func(self, model, column, key, iter, cols):
        """Equal function for interactive search"""
        search_string = ''
        for col in cols:
            search_string += ' ' + model[iter][col].lower()
        for word in key.split():
            if word.lower() not in search_string:
                return True
        return False
    
    # Callbacks
    
    def onRefreshClicked(self, button):
        """Refresh screen on button click"""
        
        # Sanitise entries
        if self.entry_top.get_text() == '':
            self.entry_top.set_text('1')
        if self.entry_bottom.get_text() == '':
            self.entry_bottom.set_text('1')
        if self.entry_left.get_text() == '':
            self.entry_left.set_text('1')
        
        if self.spreadsheet:
            self.update()
    
    def onEntryEditedNum(self, entry):
        """Treeview cell renderer for editable number field
        
            User Data:
                column: column in ListStore being edited
        """
        new_text = entry.get_text()
        num = ''
        if new_text is not '':
            try:  # check whether item evaluates fine
                num = int(new_text)
                if num <= 0:
                   num = 1
            except:
                log.warning("SpreadsheetDialog - onEntryEditedNum - evaluation of [" 
                + new_text + "] failed")
        entry.set_text(str(num))

            
class Spreadsheet:
    """Manage input and output of spreadsheets"""
    
    def __init__(self, filename=None):
        if filename is not None:
            self.spreadsheet = openpyxl.load_workbook(filename)
        else:
            self.spreadsheet = openpyxl.Workbook()
        self.sheet = self.spreadsheet.active
    
    def save(self, filename):
        """Save worksheet to file"""
        self.spreadsheet.save(filename)
        
    # Sheet management
    
    def new_sheet(self):
        """Create a new sheet to spreadsheet and set as active"""
        self.sheet = self.spreadsheet.create_sheet()  
            
    def sheets(self):
        """Returns a list of sheetnames"""
        return self.spreadsheet.get_sheet_names()
        
    def length(self):
        """Get number of rows in sheet"""
        return len(self.sheet.rows)
        
    def set_title(self, title):
        """Set title of sheet"""
        self.sheet.title = title
        
    def set_column_widths(self, widths):
        """Set column widths of sheet"""
        for column, width in enumerate(widths, 1):
            col_letter = openpyxl.cell.get_column_letter(column)
            self.sheet.column_dimensions[col_letter].width = width
        
    def set_active_sheet(self, sheetref):
        """Set active sheet of spreadsheet"""
        sheetname = ''
        sheetno = None
        if type(sheetref) is int:
            sheetno = sheetref
        elif type(sheetref) is str:
            sheetname = sheetref
        
        if sheetname in self.sheets():
            self.sheet = self.spreadsheet[sheetname]
        elif sheetno is not None and sheetno < len(self.sheets()):
            self.sheet = self.spreadsheet[self.sheets()[sheetno]]
    
    def set_style(self, row, col, bold=False, wrap_text=True, horizontal='general'):
        """Set style of individual cell"""
        font = openpyxl.styles.Font(bold=bold)
        alignment = openpyxl.styles.Alignment(wrap_text=wrap_text, horizontal=horizontal)
        self.sheet.cell(row=row, column=col).font = font
        self.sheet.cell(row=row, column=col).alignment = alignment
        
    # Data addition functions
            
    def append(self, ss_obj):
        """Append an sheet to current sheet"""
        sheet = ss_obj.spreadsheet.active
        rowcount = self.length()
        for row_no, row in enumerate(sheet.rows, 1):
            for col_no, cell in enumerate(row, 1):
                self.sheet.cell(row=row_no+rowcount, column=col_no).value = cell.value
                self.sheet.cell(row=row_no+rowcount, column=col_no).style = cell.style
                
    def append_data(self, data, bold=False, wrap_text=True, horizontal='general'):
        """Append data to current sheet"""
        rowcount = self.length()
        self.insert_data(data, rowcount+1, 1, bold, wrap_text, horizontal)
    
    def insert_data(self, data, start_row=1, start_col=1, bold=False, wrap_text=True, horizontal='general'):
        """Insert data to current sheet"""
        # Setup styles
        font = openpyxl.styles.Font(bold=bold)
        alignment = openpyxl.styles.Alignment(wrap_text=wrap_text, horizontal=horizontal)
        # Apply data and styles
        for row_no, row in enumerate(data, start_row):
            for col_no, value in enumerate(row, start_col):
                self.sheet.cell(row=row_no, column=col_no).value = value
                self.sheet.cell(row=row_no, column=col_no).font = font
                self.sheet.cell(row=row_no, column=col_no).alignment = alignment
                
    def add_merged_cell(self, value, row=None, width=2, bold=False, wrap_text=True, horizontal='general'):
        """Add a merged cell of prescrbed width"""
        if row is None:
            rowstart = self.length() + 1
        else:
            rowstart = row
        self.sheet.merge_cells(start_row=rowstart,start_column=1,end_row=rowstart,end_column=width)
        self.__setitem__([rowstart,1], value)
        self.set_style(rowstart, 1, bold, wrap_text, horizontal)
    
    def __setitem__(self, index, value):
        """Set an individual cell"""
        self.sheet.cell(row=index[0], column=index[1]).value = value
        
    def __getitem__(self, index):
        """Set an individual cell"""
        return self.sheet.cell(row=index[0], column=index[1]).value
            
    # Bulk read functions
    
    def read_rows(self, columntypes = [], start=0, end=-1, left=0):
        """Read and validate selected rows from current sheet"""
        # Get count of rows
        rowcount = self.length()
        if end < 0 or end >= rowcount:
            count_actual = rowcount
        else:
            count_actual = end
        
        items = []
        for row in range(start, count_actual):
            cells = []
            skip = 0  # No of columns to be skiped ex. breakup, total etc...
            for columntype, i in zip(columntypes, list(range(left, len(columntypes)+left))):
                cell = self.sheet.cell(row = row + 1, column = i - skip + 1).value
                if columntype == MEAS_DESC:
                    if cell is None:
                        cell_formated = ""
                    else:
                        cell_formated = str(cell)
                elif columntype == MEAS_L:
                    if cell is None:
                        cell_formated = "0"
                    else:
                        try:  # try evaluating float
                            cell_formated = str(float(cell))
                        except:
                            cell_formated = '0'
                elif columntype == MEAS_NO:
                    if cell is None:
                        cell_formated = "0"
                    else:
                        try:  # try evaluating int
                            cell_formated = str(int(cell))
                        except:
                            cell_formated = '0'
                else:
                    cell_formated = ''
                    log.warning("Spreadsheet - Value skipped on import - " + str((row, i)))
                if columntype == MEAS_CUST:
                    skip = skip + 1
                cells.append(cell_formated)
            items.append(cells)
        return items


class LatexFile:
    """Class for formating and rendering of latex code"""
    
    def __init__(self, latex_buffer = ""):
        self.latex_buffer = latex_buffer
    
    # Inbuilt methods

    def clean_latex(self, text):
        """Replace special charchters with latex commands"""
        for splchar, replspelchar in zip(['\\', '#', '$', '%', '^', '&', '_', '{', '}', '~', '\n'],
                                         ['\\textbackslash ', '\# ', '\$ ', '\% ', '\\textasciicircum ', '\& ', '\_ ',
                                          '\{ ', '\} ', '\\textasciitilde ', '\\newline ']):
            text = text.replace(splchar, replspelchar)
        return text
        
    # Operator overloading
    
    def __add__(self,other):
        return LatexFile(self.latex_buffer + '\n' + other.latex_buffer)
        
    # Public members
    
    def get_buffer(self):
        """Get underlying latex buffer"""
        return self.latex_buffer
            
    def add_preffix_from_file(self,filename):
        """Add a latex file as preffix"""
        latex_file = open(filename,'r', encoding="utf8")
        self.latex_buffer = latex_file.read() + '\n' + self.latex_buffer
        latex_file.close()
        
    def add_suffix_from_file(self,filename):
        """Add a latex file as suffix"""
        latex_file = open(filename,'r', encoding="utf8")
        self.latex_buffer = self.latex_buffer + '\n' + latex_file.read()
        latex_file.close()
        
    def replace_and_clean(self, dic):
        """Replace items as per dictionary after cleaning special charachters"""
        for i, j in dic.items():
            j = self.clean_latex(j)
            self.latex_buffer = self.latex_buffer.replace(i, j)

    def replace(self, dic):
        """Replace items as per dictionary"""
        for i, j in dic.items():
            self.latex_buffer = self.latex_buffer.replace(i, j)
            
    def write(self, filename):
        """Write latex file to disk"""
        file_latex = open(filename,'w', encoding="utf8")
        file_latex.write(self.latex_buffer)
        file_latex.close()
        
        
class ProgressWindow:
    """Class for handling display of long running proccess"""
    
    def __init__(self, parent):
        # Setup data
        self.step = 0
        self.fraction = 0
        
        # Setup progress indicator
        self.dialog = Gtk.Window(default_height=250, default_width=400, title='Process running...')
        self.dialog.set_transient_for(parent)
        self.dialog.set_gravity(Gdk.Gravity.CENTER)
        self.dialog.set_position(Gtk.WindowPosition.CENTER_ON_PARENT)
        self.dialog.set_modal(True)
        self.dialog.set_type_hint(Gdk.WindowTypeHint.DIALOG)
        box = Gtk.Box(orientation=Gtk.Orientation.VERTICAL, spacing=6)
        box.set_margin_left(6)
        box.set_margin_right(6)
        
        scroll = Gtk.ScrolledWindow()
        scroll.set_vexpand(True)
        self.store = Gtk.ListStore(str)
        self.tree = Gtk.TreeView.new_with_model(self.store)
        column = Gtk.TreeViewColumn("Message")
        cell = Gtk.CellRendererText()
        column.pack_start(cell, True)
        column.add_attribute(cell, "markup", 0)
        self.tree.append_column(column)
        scroll.add(self.tree)
        
        box.pack_start(scroll, True, True, 3)
        self.progress = Gtk.ProgressBar()
        box.pack_start(self.progress, False, False, 3)
        dismiss = Gtk.Button('Dismiss')
        box.pack_start(dismiss, False, False, 3)
        self.dialog.add(box)
        
        # Connect events
        self.dialog.connect("delete-event",self.on_delete)
        dismiss.connect("clicked",self.on_dismiss)
        
    def show(self):
        def callback():
            self.dialog.show_all()
            return False
            
        GLib.idle_add(callback)
    
    def close(self):
        self.dialog.close()
        
    def on_delete(self, *args):
        if self.fraction == 1:
            return False
        else:
            # Cancel event
            self.dialog.hide()
            return True
    
    def on_dismiss(self, button):
        if self.fraction == 1:
            self.close()
        else:
            self.dialog.hide()
        
    def set_pulse_step(self, width):
        self.step = width
        self.fraction = 0
        
    def pulse(self, end=False):
        def callback():
            self.fraction += self.step
            if end:
                self.fraction = 1
                self.dialog.set_title('Process Complete')
                self.show()
            self.progress.set_fraction(self.fraction)
            return False
        GLib.idle_add(callback)
        
    def add_message(self, message):
        def callback():
            itemiter = self.store.append([message])
            path = self.store.get_path(itemiter)
            self.tree.scroll_to_cell(path)
            return False
        GLib.idle_add(callback)
        

class Command(object):
    """Runs a command in a seperate thread"""
    
    def __init__(self, cmd):
        """Initialises class with command to be executed"""
        self.cmd = cmd
        self.process = None

    def run(self, timeout):
        """Run set command with selected timeout"""
        def target():
            if platform.system() == 'Linux':
                self.process = subprocess.Popen(self.cmd)
            elif platform.system() == 'Windows':
                self.process = subprocess.Popen(self.cmd, shell=True)
            log.info('Sub-process spawned - ' + str(self.process.pid))
            self.process.communicate()
        thread = threading.Thread(target=target)
        thread.start()

        thread.join(timeout)
        if thread.is_alive():
            log.error('Terminating sub-process exceeding timeout - ' + str(self.process.pid))
            self.process.terminate()
            thread.join()
            return -1
        return 0


## GLOBAL METHODS

def get_user_input_text(parent, message, title='', oldval=None, multiline=False):
    '''Gets a single user input by diplaying a dialog box
    
    Arguments:
        parent: Parent window
        message: Message to be displayed to user
        title: Dialog title text
        multiline: Allows multiline input is True
    Returns:
        Returns user input as a string or 'None' if user does not input text.
    '''
    dialogWindow = Gtk.MessageDialog(parent,
                                     Gtk.DialogFlags.MODAL | Gtk.DialogFlags.DESTROY_WITH_PARENT,
                                     Gtk.MessageType.QUESTION,
                                     Gtk.ButtonsType.OK_CANCEL,
                                     message)

    dialogWindow.set_transient_for(parent)
    dialogWindow.set_title(title)
    dialogWindow.set_default_response(Gtk.ResponseType.OK)

    dialogBox = dialogWindow.get_content_area()
    text = ''
    
    if multiline:
        # Function to mark first line as bold
        def mark_heading(textbuff, tag):
            start = textbuff.get_start_iter()
            end = textbuff.get_end_iter()
            textbuff.remove_all_tags(start, end)
            match = start.forward_search('\n', 0, end)
            if match != None:
                match_start, match_end = match
                textbuff.apply_tag(tag, start, match_start)

        scrolledwindow = Gtk.ScrolledWindow()
        scrolledwindow.set_hexpand(True)
        scrolledwindow.set_vexpand(True)
        scrolledwindow.set_size_request(300, 100)

        textview = Gtk.TextView()
        textbuffer = textview.get_buffer()
        dialogBox.pack_end(scrolledwindow, False, False, 0)
        scrolledwindow.add(textview)
        
        # Set old value
        if oldval != None:
            textbuffer.set_text(oldval)
        
        # Mark heading
        tag_bold = textbuffer.create_tag("bold", weight=Pango.Weight.BOLD)
        mark_heading(textbuffer, tag_bold)
        textbuffer.connect("changed", mark_heading, tag_bold)

        dialogWindow.show_all()
        response = dialogWindow.run()
        text = textbuffer.get_text(textbuffer.get_start_iter(),textbuffer.get_end_iter(), True)
    else:
        userEntry = Gtk.Entry()
        userEntry.set_activates_default(True)
        userEntry.set_size_request(50, 0)
        dialogBox.pack_end(userEntry, False, False, 0)
        
        # Set old value
        if oldval != None:
            userEntry.set_text(oldval)

        dialogWindow.show_all()
        response = dialogWindow.run()
        text = userEntry.get_text()
    dialogWindow.destroy()
    if (response == Gtk.ResponseType.OK) and (text != ''):
        return text
    else:
        return None

def abs_path(*args):
    """Returns absolute path to the relative path provided"""
    return os.path.join(os.path.split(__file__)[0],*args)

def posix_path(*args):
    """Returns platform independent filename"""
    if platform.system() == 'Linux': 
        if len(args) > 1:
            return posixpath.join(*args)
        else:
            return args[0]
    elif platform.system() == 'Windows':
        if len(args) > 1:
            path = os.path.normpath(posixpath.join(*args))
        else:
            path = os.path.normpath(args[0])
        # remove any leading slash
        if path[0] == '\\':
            return path[1:]
        else:
            return path
            
def open_file(filename):
    if platform.system() == 'Linux':
        subprocess.call(('xdg-open', abs_path(filename)))
    elif platform.system() == 'Windows':
        os.startfile(abs_path(filename))
        
def get_file_path_from_dnd_dropped_uri(uri):
    # Get the path to file
    path = ""
    if uri.startswith('file:\\\\\\'): # windows
        path = uri[8:] # 8 is len('file:///')
    elif uri.startswith('file://'): # nautilus, rox
        path = uri[7:] # 7 is len('file://')
    elif uri.startswith('file:'): # xffm
        path = uri[5:] # 5 is len('file:')

    path = url2pathname(path) # escape special chars
    path = path.strip('\r\n\x00') # remove \r\n and NULL

    return path
            
def run_latex(folder, filename): 
    """Runs latex on file to folder in two passes"""
    if filename is not None:
        latex_exec = Command([global_settings_dict['latex_path'], '-interaction=batchmode', '-output-directory=' + folder, filename])
        # First Pass
        code = latex_exec.run(timeout=LATEX_TIMEOUT)
        if code == 0:
            # Second Pass
            code = latex_exec.run(timeout=LATEX_TIMEOUT)
            if code != 0:
                return CMB_ERROR
        else:
            return CMB_ERROR
    return CMB_OK

def clean_markup(text):
    """Clear markup text of special characters"""
    for splchar, replspelchar in zip(['&', '<', '>', ], ['&amp;', '&lt;', '&gt;']):
        text = text.replace(splchar, replspelchar)
    return text
    