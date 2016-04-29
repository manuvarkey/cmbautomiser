#!/usr/bin/env python
# -*- coding: utf-8 -*-
#
# schedule_dialog.py
#
#  Copyright 2014 Manu Varkey <manuvarkey@gmail.com>
#  
#  This program is free software; you can redistribute it and/or modify
#  it under the terms of the GNU General Public License as published by
#  the Free Software Foundation; either version 2 of the License, or
#  (at your option) any later version.
#  
#  This program is distributed in the hope that it will be useful,
#  but WITHOUT ANY WARRANTY; without even the implied warranty of
#  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#  GNU General Public License for more details.
#  
#  You should have received a copy of the GNU General Public License
#  along with this program; if not, write to the Free Software
#  Foundation, Inc., 51 Franklin Street, Fifth Floor, Boston,
#  MA 02110-1301, USA.
#  
#  

import copy, logging

from gi.repository import Gtk, Gdk, GLib

from undo import *

# local files import
from schedule import *
from misc import *
from openpyxl import Workbook, load_workbook

# Setup logger object
log = logging.getLogger(__name__)

# class storing individual items in schedule of work
class ScheduleItemGeneric:
    def __init__(self, item=None):
        self.item = item

    def set(self, item):
        self.item = item

    def get(self):
        return self.item

    def __setitem__(self, index, value):
        self.item[index] = value

    def __getitem__(self, index):
        return self.item[index]

    def print_item(self):
        print(self.item)


class ScheduleGeneric:
    def __init__(self):
        self.items = []  # main data store of rows

    def append_item(self, item):
        self.items.append(item)

    def get_item_by_index(self, index):
        return self.items[index]

    def set_item_at_index(self, index, item):
        self.items[index] = item

    def insert_item_at_index(self, index, item):
        self.items.insert(index, item)

    def remove_item_at_index(self, index):
        del (self.items[index])

    def __setitem__(self, index, value):
        self.items[index] = value

    def __getitem__(self, index):
        return self.items[index]

    def length(self):
        return len(self.items)

    def clear(self):
        del self.items[:]

    def print_item(self):
        print("schedule start")
        for item in self.items:
            item.print_item()
        print("schedule end")


class ScheduleDialog:
    # General Methods

    def get_user_input_text(self, parent, message, title='', value=None):
        # Returns user input as a string or None
        # If user does not input text it returns None, NOT AN EMPTY STRING.
        dialog_window = Gtk.MessageDialog(parent,
                                         Gtk.DialogFlags.MODAL | Gtk.DialogFlags.DESTROY_WITH_PARENT,
                                         Gtk.MessageType.QUESTION,
                                         Gtk.ButtonsType.OK_CANCEL,
                                         message)

        dialog_window.set_title(title)
        dialog_window.set_default_response(Gtk.ResponseType.OK)

        dialogBox = dialog_window.get_content_area()
        userEntry = Gtk.Entry()
        userEntry.set_activates_default(True)
        userEntry.set_size_request(100, 0)
        dialogBox.pack_end(userEntry, False, False, 0)

        if value is not None:
            userEntry.set_text(value)

        dialog_window.show_all()
        response = dialog_window.run()
        
        if (response == Gtk.ResponseType.OK):
            text = userEntry.get_text()
            dialog_window.destroy()
            return text
        else:
            dialog_window.destroy()
            return None

    # return packed combo box
    def get_item_schedule_combobox(self):

        # combo box
        combo = Gtk.ComboBox.new_with_model(self.item_schedule_store)

        # text entries to be packed
        combo_text_agmntno = Gtk.CellRendererText()
        combo_divider1 = Gtk.CellRendererText()
        combo_text_desc = Gtk.CellRendererText()
        combo_divider2 = Gtk.CellRendererText()
        combo_text_unit = Gtk.CellRendererText()

        # start packing
        combo.pack_start(combo_text_agmntno, False)
        combo.pack_start(combo_divider1, False)
        combo.pack_start(combo_text_desc, True)
        combo.pack_start(combo_divider2, False)
        combo.pack_start(combo_text_unit, False)

        # set attributes
        combo.add_attribute(combo_text_agmntno, 'text', 0)
        combo.add_attribute(combo_text_desc, 'text', 1)
        combo.add_attribute(combo_text_unit, 'text', 2)

        combo.props.id_column = 0

        combo_divider1.props.text = '|'
        combo_divider2.props.text = '|'
        combo_text_desc.props.max_width_chars = CMB_DESCRIPTION_WIDTH
        combo_text_desc.props.wrap_width = CMB_DESCRIPTION_WIDTH
        combo_text_desc.props.ellipsize = 2
        combo_text_desc.set_fixed_size(-1, 5)
        combo_text_desc.set_fixed_height_from_font(1)

        return combo

    # General signal handler Methods

    # for selecting item schedule
    def onComboChanged(self, combo, index):
        tree_iter = combo.get_active_iter()
        if tree_iter is not None:
            model = combo.get_model()
            itemno = model[tree_iter][0]
            self.itemnos[index] = itemno
            self.items[index] = self.item_schedule.get_item(itemno)
        else:
            self.itemnos[index] = None
            self.items[index] = None

    def onButtonScheduleAddPressed(self, button):
        items = []
        items.append(ScheduleItemGeneric([""] * self.model_width()))
        self.insert_item_at_selection(items)

    def onButtonScheduleAddMultPressed(self, button):
        toplevel = self.tree.get_toplevel()  # get current top level window
        userInput = self.get_user_input_text(self.window, "Please enter the number \nof rows to be inserted",
                                             "Number of Rows")
        try:
            number_of_rows = int(userInput)
        except:
            log.warning("Invalid number of rows specified")
            return
        items = []
        for i in range(0, number_of_rows):
            items.append(ScheduleItemGeneric([""] * self.model_width()))
        self.insert_item_at_selection(items)

    def onButtonScheduleDeletePressed(self, button):
        self.delete_selection()

    def onUndoSchedule(self, button):
        self.undo()

    def onRedoSchedule(self, button):
        self.redo()

    def onCopySchedule(self, button):
        self.copy_selection()

    def onPasteSchedule(self, button):
        self.paste_at_selection()


    def onImportScheduleClicked(self, button):
        filename = self.builder.get_object("filechooserbutton_schedule").get_filename()
        spreadsheet = load_workbook(filename)
        sheet = spreadsheet.active
        # get count of rows
        rowcount = len(sheet.columns)
        items = []
        for row in range(1, rowcount):
            cells = []
            skip = 0  # no of columns to be skiped ex. breakup, total etc...
            for column_type, i in zip(self.columntypes, list(range(len(self.columntypes)))):
                cell = sheet.cell(row = row + 1, column = i - skip + 1).value
                if cell is None:
                    cell_formated = ""
                else:
                    try:  # try evaluating string
                        if column_type == MEAS_DESC:
                            cell_formated = str(cell)
                        elif column_type == MEAS_L:
                            cell_formated = str(float(cell))
                        elif column_type == MEAS_NO:
                            cell_formated = str(int(cell))
                        else:
                            cell_formated = ''
                    except:
                        cell_formated = ''
                if column_type == MEAS_CUST:
                    skip = skip + 1
                cells.append(cell_formated)
            item = ScheduleItemGeneric(cells)
            items.append(item)

        self.insert_item_at_selection(items)

    # Cell renderer treeview generic for editable text field
    def onScheduleCellEditedText(self, widget, row, new_text, column):
        self.cell_renderer_text(int(row), column, new_text)

    # Cell renderer treeview generic for editable number field
    def onScheduleCellEditedNum(self, widget, row, new_text, column):
        try:  # check whether item evaluates fine
            eval(new_text)
        except:
            return
        self.cell_renderer_text(int(row), column, new_text)

    def onEditStarted(self, widget, editable, path, column):
        row = int(path)
        item = self.schedule.get_item_by_index(row).get()
        editable.props.text = item[column]

    # for browsing with tab key
    def onKeyPressTreeviewSchedule(self, treeview, event):
        keyname = event.get_keyval()[1]
        state = event.get_state()
        shift_pressed = bool(state & Gdk.ModifierType.SHIFT_MASK)
        path, col = treeview.get_cursor()
        if path != None:
            ## only visible columns!!
            columns = [c for c in treeview.get_columns() if c.get_visible() and self.celldict[c].props.editable]
            rows = [r for r in treeview.get_model()]
            if col in columns:
                colnum = columns.index(col)
                rownum = path[0]
                if keyname in [Gdk.KEY_Tab, Gdk.KEY_ISO_Left_Tab]:
                    if shift_pressed == 1:
                        if colnum - 1 >= 0:
                            prev_column = columns[colnum - 1]
                        else:
                            tmodel = treeview.get_model()
                            titer = tmodel.iter_previous(tmodel.get_iter(path))
                            if titer is None:
                                titer = tmodel.get_iter_first()
                                path = tmodel.get_path(titer)
                                prev_column = columns[0]
                            else:
                                path = tmodel.get_path(titer)
                                prev_column = columns[-1]
                        GLib.timeout_add(50, treeview.set_cursor, path, prev_column, True)
                    else:
                        if colnum + 1 < len(columns):
                            next_column = columns[colnum + 1]
                        else:
                            tmodel = treeview.get_model()
                            titer = tmodel.iter_next(tmodel.get_iter(path))
                            if titer is None:
                                titer = tmodel.get_iter_first()
                            path = tmodel.get_path(titer)
                            next_column = columns[0]
                        GLib.timeout_add(50, treeview.set_cursor, path, next_column, True)
                elif keyname in [Gdk.KEY_Return, Gdk.KEY_KP_Enter]:
                    if shift_pressed == 1:
                        if rownum - 1 >= 0:
                            rownum -= 1
                        else:
                            rownum = 0
                        path = [rownum]
                        GLib.timeout_add(50, treeview.set_cursor, path, col, True)
                    else:
                        if rownum + 1 < len(rows):
                            rownum += 1
                        else:
                            rownum = len(rows) - 1
                        path = [rownum]
                        GLib.timeout_add(50, treeview.set_cursor, path, col, True)
                elif keyname in [Gdk.KEY_Control_L,Gdk.KEY_Control_R]:  # unselect all
                    self.tree.get_selection().unselect_all()

    def onClearButtonPressed(self, button, combo):
        combo.set_active(-1)
        self.onComboChanged(combo, -1)

    # Class methods

    @undoable
    def append_item(self, itemlist):
        newrows = []

        for item in itemlist:
            self.schedule.append_item(item)
            newrows.append(self.schedule.length() - 1)
        self.update_store()

        yield "Append data items to schedule at row '{}'".format(newrows)
        # Undo action
        self.delete_row(newrows)
        self.update_store()

    def insert_item_at_selection(self, itemlist):
        selection = self.tree.get_selection()
        if selection.count_selected_rows() != 0:  # if selection exists
            [model, paths] = selection.get_selected_rows()
            rows = []
            for i in range(0, len(itemlist)):
                rows.append(paths[0].get_indices()[0] + 1)
            self.insert_item_at_row(itemlist, rows)
        else:  # if no selection
            self.append_item(itemlist)

    @undoable
    def insert_item_at_row(self, itemlist, rows):  # note needs rows to be sorted
        newrows = []
        for i in range(0, len(rows)):
            self.schedule.insert_item_at_index(rows[i] + i - 1, itemlist[i])
            newrows.append(rows[i] + i - 1)
        self.update_store()

        yield "Insert data items to schedule at rows '{}'".format(rows)
        # Undo action
        self.delete_row(newrows)
        self.update_store()

    def delete_selection(self):
        # get rows to delete
        selection = self.tree.get_selection()
        [model, paths] = selection.get_selected_rows()
        rows = []
        for path in paths:  # get rows
            rows.append(int(path.get_indices()[0]))

        # delete rows
        self.delete_row(rows)

    @undoable
    def delete_row(self, rows):
        newrows = []
        items = []

        rows.sort()
        for i in range(0, len(rows)):
            items.append(self.schedule[rows[i] - i])
            newrows.append(rows[i] + i + 1)
            self.schedule.remove_item_at_index(rows[i] - i)
        self.update_store()

        yield "Delete data items from schedule at rows '{}'".format(rows)
        # Undo action
        self.insert_item_at_row(items, newrows)
        self.update_store()

    # For undoable function for modifying value of a cell
    @undoable
    def cell_renderer_text(self, row, column, newvalue):
        oldvalue = self.schedule[row][column]
        self.schedule[row][column] = newvalue
        self.update_store()

        yield "Change data item at row:'{}' and column:'{}'".format(row, column)
        # Undo action
        self.schedule[row][column] = oldvalue
        self.update_store()

    def copy_selection(self):
        selection = self.tree.get_selection()
        if selection.count_selected_rows() != 0:  # if selection exists
            [model, paths] = selection.get_selected_rows()
            items = []
            for path in paths:
                row = int(path.get_indices()[0])  # get item row
                item = self.schedule[row]
                items.append(item)  # save items
            text = pickle.dumps(items)  # dump item as text
            self.clipboard.set_text(text, -1)  # push to clipboard
        else:  # if no selection
            log.warning("No items selected to copy")

    def paste_at_selection(self):
        text = self.clipboard.wait_for_text()  # get text from clipboard
        if text is not None:
            try:
                itemlist = pickle.loads(text)  # recover item from string
                if isinstance(itemlist[0], ScheduleItemGeneric) and len(itemlist[0].get()) == self.model_width():
                    selection = self.tree.get_selection()
                    if selection.count_selected_rows() != 0:  # if selection exists
                        [model, paths] = selection.get_selected_rows()
                        rows = []
                        for i in range(0, len(itemlist)):
                            rows.append(int(paths[0].get_indices()[0] + 1))
                        self.insert_item_at_row(itemlist, rows)
                    else:  # if no selection
                        self.append_item(itemlist)
            except:
                log.warning('No valid data in clipboard')
        else:
            log.warning("No text on the clipboard.")

    def model_width(self):
        return len(self.columntypes)

    def get_model(self):
        remark = self.remark_cell.get_text()
        item_remarks = [cell.get_text() for cell in self.item_remarks_cell]
        records = []
        for record in self.schedule:
            records.append(record.get())
        data = [self.itemnos, self.items, records, remark, item_remarks]
        return data

    def set_model(self, data):
        self.itemnos = data[0]
        self.items = data[1]

        for itemno, combo in zip(self.itemnos, self.item_combos):
            combo.set_active_id(itemno)

        self.schedule.clear()
        for record in data[2]:
            self.schedule.append_item(copy.deepcopy(ScheduleItemGeneric(record)))
        self.remark_cell.set_text(data[3])
        for cell, text in zip(self.item_remarks_cell, data[4]):
            cell.set_text(text)
        self.update_store()

    def undo(self):
        setstack(self.stack)  # select schedule undo stack
        log.info(self.stack.undotext())
        self.stack.undo()

    def redo(self):
        setstack(self.stack)  # select schedule undo stack
        log.info(self.stack.redotext())
        self.stack.redo()

    def clear(self):
        self.stack.clear()
        self.schedule.clear()
        self.update_store()

    def update_store(self):
        # Add or remove required rows
        rownum = 0
        for row in self.store:
            rownum += 1
        if rownum > self.schedule.length():
            for i in range(rownum - self.schedule.length()):
                del self.store[-1]
        else:
            for i in range(self.schedule.length() - rownum):
                self.store.append()

        # Find formated items and fill in values
        for row in range(0, self.schedule.length()):
            item = self.schedule.get_item_by_index(row).get()
            display_item = []
            for item_elem, columntype, render_func in zip(item, self.columntypes, self.render_funcs):
                try:
                    if item_elem != "" or columntype == MEAS_CUST:
                        if columntype == MEAS_CUST:
                            display_item.append(render_func(item, row))
                        if columntype == MEAS_DESC:
                            display_item.append(item_elem)
                        elif columntype == MEAS_NO:
                            display_item.append(str(int(round(eval(item_elem)))))
                        elif columntype == MEAS_L:
                            value = str(round(eval(item_elem), 2)) if item_elem != '0' else ''
                            display_item.append(value)
                    else:
                        display_item.append("")
                except TypeError:
                    display_item.append("")
                    log.warning('Error: Wrong value loaded in store')
            self.store[row] = display_item
        setstack(self.stack)  # select undo stack

    def __init__(self, parent, itemnos, captions, columntypes, render_funcs, item_schedule):

        # Setup variables
        self.parent = parent
        self.itemnos = itemnos
        self.captions = captions
        self.columntypes = columntypes
        self.render_funcs = render_funcs
        self.item_schedule = item_schedule
        self.items = [self.item_schedule.get_item(itemno) for itemno in self.itemnos]

        self.schedule = ScheduleGeneric()

        # Setup dialog window
        self.builder = Gtk.Builder()
        self.builder.add_from_file(abs_path("interface","scheduledialog.glade"))
        self.window = self.builder.get_object("dialog")
        self.window.set_transient_for(self.parent)
        self.builder.connect_signals(self)

        # Get required objects

        self.listbox_itemnos = self.builder.get_object("listbox_itemnos")
        self.tree = self.builder.get_object("treeview_schedule")

        # Setup liststore model for schedule

        self.item_schedule_store = Gtk.ListStore(str, str, str, float, str)
        for row in range(self.item_schedule.length()):
            item = self.item_schedule.get_item_by_index(row)
            self.item_schedule_store.append([item.itemno, item.description, item.unit, item.rate, item.reference])

        # setup remarks row
        row = Gtk.ListBoxRow()
        hbox = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL)
        label = Gtk.Label("Remarks:")
        entry = Gtk.Entry()

        # pack row
        row.add(hbox)
        hbox.pack_start(label, False, True, 3)
        hbox.pack_start(entry, True, True, 3)

        # set additional properties
        label.props.width_request = 50

        # add to list box
        self.listbox_itemnos.add(row)

        self.remark_cell = entry
        self.item_combos = []
        self.item_remarks_cell = []

        for itemno, index in zip(self.itemnos, list(range(len(self.itemnos)))):
            # get items in row
            row = Gtk.ListBoxRow()
            hbox = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL)
            entry = Gtk.Entry()
            button = Gtk.Button(stock=Gtk.STOCK_CLEAR)
            combo = self.get_item_schedule_combobox()

            # pack row
            row.add(hbox)
            hbox.pack_start(entry, False, True, 3)
            hbox.pack_start(combo, True, True, 3)
            hbox.pack_start(button, False, True, 0)

            # set additional properties
            entry.props.width_request = 50
            button.connect("clicked", self.onClearButtonPressed, combo)
            combo.connect("changed", self.onComboChanged, index)

            # add to list box
            self.listbox_itemnos.add(row)

            # save variables
            self.item_combos.append(combo)
            self.item_remarks_cell.append(entry)

        # Setup treeview
        data_types = [str] * len(self.columntypes)
        self.store = Gtk.ListStore(*data_types)
        self.tree.set_model(self.store)

        self.celldict = {}

        for columntype, caption, render_func, i in zip(self.columntypes, self.captions, self.render_funcs,
                                                       list(range(len(self.columntypes)))):
            cell = Gtk.CellRendererText()

            column = Gtk.TreeViewColumn(caption, cell, text=i)
            column.props.resizable = True

            self.celldict[column] = cell  # add cell to column map for future ref

            self.tree.append_column(column)
            self.tree.props.search_column = 0

            if columntype == MEAS_NO:
                cell.set_property("editable", True)
                column.props.min_width = 50
                column.props.fixed_width = 75
                if render_func is None:
                    cell.connect("edited", self.onScheduleCellEditedNum, i)
                    cell.connect("editing_started", self.onEditStarted, i)
                else:
                    cell.connect("edited", render_func, i)
            elif columntype == MEAS_L:
                cell.set_property("editable", True)
                column.props.min_width = 50
                column.props.fixed_width = 75
                if render_func is None:
                    cell.connect("edited", self.onScheduleCellEditedNum, i)
                    cell.connect("editing_started", self.onEditStarted, i)
                else:
                    cell.connect("edited", render_func, i)
            elif columntype == MEAS_DESC:
                cell.set_property("editable", True)
                column.props.fixed_width = 200
                column.props.min_width = 100
                column.props.expand = True
                cell.props.wrap_width = 500
                cell.props.wrap_mode = 2
                if render_func is None:
                    cell.connect("edited", self.onScheduleCellEditedText, i)
                    cell.connect("editing_started", self.onEditStarted, i)
                else:
                    cell.connect("edited", render_func, i)
            elif columntype == MEAS_CUST:
                cell.set_property("editable", False)
                # column.props.fixed_width = 100
                column.props.min_width = 50
                cell.props.wrap_width = 200
                cell.props.wrap_mode = 2
                if render_func is None:
                    cell.connect("edited", self.onScheduleCellEditedText, i)
                    cell.connect("editing_started", self.onEditStarted, i)
                else:
                    cell.connect("edited", render_func, i)

        # intialise undo redo stack
        self.stack = Stack()  # initialise undo/redo stack
        setstack(self.stack)  # select schedule undo stack

        # intialise clipboard
        self.clipboard = Gtk.Clipboard.get(Gdk.SELECTION_CLIPBOARD)  # initialise clipboard

        # Connect signals with custom userdata

        self.tree.connect("key-press-event", self.onKeyPressTreeviewSchedule)

    def run(self):
        self.window.show_all()
        response = self.window.run()

        if response == 1:
            data = self.get_model()
            self.window.destroy()
            return data
        else:
            self.window.destroy()
            return None
