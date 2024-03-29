#!/usr/bin/env python
# -*- coding: utf-8 -*-
#
# bill.py
#
#  Copyright 2014 Manu Varkey <manuvarkey@gmail.com>
#
#  This program is free software; you can redistribute it and/or modify
#  it under the terms of the GNU General Public License as published by
#  the Free Software Foundation; either version 2 of the License, or
#  (at your option) any later version.
#
#  This program is distributed in the hope that it will be useful,
#  but WITHOUT ANY WARRANTY; without even the implied warranty of
#  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#  GNU General Public License for more details.
#
#  You should have received a copy of the GNU General Public License
#  along with this program; if not, write to the Free Software
#  Foundation, Inc., 51 Franklin Street, Fifth Floor, Boston,
#  MA 02110-1301, USA.
#
#

from gi.repository import Gtk, Gdk, GLib
import copy, math, logging, hashlib

from decimal import Decimal, ROUND_HALF_UP
def Currency(x):
  return Decimal(x).quantize(Decimal(".01"), rounding=ROUND_HALF_UP)

def CurrencyR(x):
  return Decimal(x).quantize(Decimal("1"), rounding=ROUND_HALF_UP)

from openpyxl import Workbook, load_workbook, worksheet
from openpyxl.styles import Alignment, Font, Border
from openpyxl.utils import get_column_letter

# local files import
from .. import misc
from . import measurement

# Setup logger object
log = logging.getLogger(__name__)

class BillData:
    """Data structure for Bill Class"""
    
    def __init__(self, bill_type=misc.BILL_NORMAL):
        self.prev_bill = None
        self.cmb_name = ''
        self.title = ''
        self.bill_date = ''
        self.starting_page = 1
        self.mitems = []  # for storing paths to billed measurement items [cmb][measurement][Meas Item]
        self.item_part_percentage = dict()  # part rate for exess rate items
        self.item_excess_part_percentage = dict()  # part rate for exess rate items
        self.item_excess_rates = dict()  # list of excess rates above excess_percentage
        self.adjustments = []

        # Additional elements for custom bill
        self.item_qty = dict()  # qtys of items b/f
        self.item_normal_amount = dict()  # total item amount for qty at normal rate
        self.item_excess_amount = dict()  # amounts for qty at excess rate
        
        # Type Flag
        self.bill_type = bill_type
        self.bill_text = ''
    
    def get_model(self):
        """Get data model"""
        model = [self.prev_bill, self.cmb_name, self.title, self.bill_date, self.starting_page,
                 self.mitems, self.item_part_percentage, self.item_excess_part_percentage ,
                 self.item_excess_rates, self.item_qty, self.item_normal_amount, self.item_excess_amount,
                 self.bill_type, self.bill_text, self.adjustments]
        return ['BillData', copy.deepcopy(model)]
    
    def set_model(self, model):
        """Set data model"""
        if model[0] == 'BillData':
            self.prev_bill = model[1][0]
            self.cmb_name = model[1][1]
            self.title = model[1][2]
            self.bill_date = model[1][3]
            self.starting_page = model[1][4]
            self.mitems = model[1][5]
            self.item_part_percentage = model[1][6]
            self.item_excess_part_percentage = model[1][7]
            self.item_excess_rates = model[1][8]
            # Additional elements for custom bill
            self.item_qty = model[1][9]
            self.item_normal_amount = model[1][10]
            self.item_excess_amount = model[1][11]
            # Type Flag
            self.bill_type = model[1][12]
            
            # HACK for supporting bill text field to be backward compatible
            if len(model[1]) > 13:
                self.bill_text = model[1][13]
            else:
                self.bill_text = ''
            # HACK for supporting adjustments field to be backward compatible
            if len(model[1]) > 14:
                self.adjustments = model[1][14]
            else:
                self.adjustments = []


class Bill:
    """Class for storing bill of work"""
    
    def __init__(self, model=None):
        self.data = BillData()
        if model is not None:
            self.data.set_model(model)
        # Derived data
        self.item_cmb_ref = dict()  # cmb references of items b/f
        self.item_paths = dict()
        self.item_qty = dict()  # qtys of items b/f
        self.item_normal_qty = dict()  # list of excess qty above excess_percentage
        self.item_normal_amount = dict()  # total item amount for qty at normal rate
        self.item_excess_qty = dict()  # qty at excess rate
        self.item_excess_amount = dict()  # amounts for qty at excess rate
        self.item_plusminus_amount = dict() # amounts corresponding to plusminus items

        self.prev_bill = None  # prev_bill data object
        self.cmb_ref = set()  # set containing refered cmbs
        self.bill_total_amount = 0  # total amount of work done uptodate
        self.bill_total_amount_plus_minus = 0  # total amount of work done on which percentage applicable
        self.bill_plusminus_amount = 0 # amount corresponding to plusminus items
        self.bill_nettotal_amount = 0 # net amount of work done uptodate after plusminus
        self.bill_since_prev_amount = 0  # since previous amount of work done
        self.bill_netpayable_amount = 0  # Net payable amount after outside work adjustments

    def clear(self, clear_all = False):
        """Clear bill data
        
            Arguments:
                clear_all: Takes values
                             True: Clears all data
                             False: Clears only derived data
        """
        if clear_all:
            self.data = BillData(self.data.bill_type)
        # Derived data
        self.item_cmb_ref = dict()  # cmb references of items b/f
        self.item_paths = dict()
        self.item_qty = dict()  # qtys of items b/f
        self.item_normal_qty = dict()  # list of excess qty above excess_percentage
        self.item_normal_amount = dict()  # total item amount for qty at normal rate
        self.item_excess_qty = dict()  # qty at excess rate
        self.item_excess_amount = dict()  # amounts for qty at excess rate
        self.item_plusminus_amount = dict() # amounts corresponding to plusminus items

        self.prev_bill = None  # prev_bill data object
        self.cmb_ref = set()  # set containing refered cmbs
        self.bill_total_amount = 0  # total amount of work done uptodate
        self.bill_total_amount_plus_minus = 0  # total amount of work done on which percentage applicable
        self.bill_plusminus_amount = 0 # component corresponding to plusminus items
        self.bill_nettotal_amount = 0 # net amount of work done uptodate after plusminus
        self.bill_since_prev_amount = 0  # since previous amount of work done
        self.bill_netpayable_amount = 0  # Net payable amount after outside work adjustments

    def get_model(self):
        """Return base model"""
        return self.data.get_model()

    def set_model(self, model):
        """Set base model"""
        self.data.set_model(model)
    
    def get_billed_items(self):
        """Return paths to billed items"""
        return self.data.mitems

    def update(self, schedule, cmbs, bills, percentage=0):
        """Update bill data structures from other objects"""
        
        # Get required datas
        itemnos = schedule.get_itemnos()
        # Clear all derived structures
        self.clear()
        
        # If bill is a normal bill
        if self.data.bill_type in (misc.BILL_NORMAL, misc.BILL_FINAL):
            # Set previous bill object
            if self.data.prev_bill is not None:
                self.prev_bill = bills[self.data.prev_bill]  # Get prev_bill object
            else:
                self.prev_bill = None
            # Initialise bill dictionaries
            for itemno in itemnos:
                self.item_cmb_ref[itemno] = []
                self.item_paths[itemno] = []
                self.item_qty[itemno] = []
                if itemno not in self.data.item_part_percentage:
                    self.data.item_part_percentage[itemno] = 100
                    self.data.item_excess_part_percentage[itemno] = 100
                    self.data.item_excess_rates[itemno] = 0

            # Fill in values from Prev Bill
            if self.prev_bill is not None:
                for itemno in itemnos:
                    if itemno in self.prev_bill.item_qty:
                        item_qty = sum(self.prev_bill.item_qty[itemno])
                        if item_qty != 0:
                            self.item_cmb_ref[itemno].append(-1)  # use -1 as marker for prev abstract
                            self.item_paths[itemno].append([self.data.prev_bill, itemno])
                            self.item_qty[itemno].append(item_qty)  # add total qty from previous bill
            # Fill in values from measurement items
            for mitem in self.data.mitems:
                item = cmbs[mitem[0]][mitem[1]][mitem[2]]
                if not isinstance(item, measurement.MeasurementItemHeading):
                    for count, (itemno, item_qty) in enumerate(zip(item.itemnos, item.get_total())):
                        # Only add if itemno is valid
                        if itemno in itemnos:
                            self.item_cmb_ref[itemno].append(mitem[0])
                            self.item_paths[itemno].append(mitem + [count])
                            self.item_qty[itemno].append(item_qty)
                        elif itemno is not None:
                            log.warning('Bill - Item No ' + str([itemno, mitem]) + ' not found, not updating in bill')

            # Evaluate remaining variables from above data
            for itemno in itemnos:
                # If item measured, calculate values
                item = schedule[itemno]
                # Determine total qty
                total_qty = sum(self.item_qty[itemno])
                # Determine rates
                normal_rate = Currency(self.data.item_part_percentage[itemno] * 0.01 * schedule[itemno].rate)
                excess_rate = Currency(self.data.item_excess_part_percentage[itemno] * 0.01 * self.data.item_excess_rates[itemno])
                # Determine items above and at normal rates
                if total_qty > (item.qty * (1 + 0.01 * item.excess_rate_percent)):
                    if misc.is_unit_item(item.unit):
                        self.item_normal_qty[itemno] = math.floor(item.qty * (1 + 0.01 * item.excess_rate_percent))
                    else:
                        self.item_normal_qty[itemno] = round(item.qty * (1 + 0.01 * item.excess_rate_percent),3)
                    self.item_excess_qty[itemno] = total_qty - self.item_normal_qty[itemno]
                else:
                    self.item_normal_qty[itemno] = total_qty
                    self.item_excess_qty[itemno] = 0
                # Determine amounts
                self.item_normal_amount[itemno] = Currency(Decimal(self.item_normal_qty[itemno]) * normal_rate)
                self.item_excess_amount[itemno] = Currency(Decimal(self.item_excess_qty[itemno]) * excess_rate)
                self.item_plusminus_amount[itemno] = Currency(self.item_normal_amount[itemno] if item.percentage else 0)
                # Update cmbs refered to by the bill
                self.cmb_ref = self.cmb_ref | set(self.item_cmb_ref[itemno])  # Add any unique cmb (find union)

            # Evaluate totals
            self.bill_total_amount = Currency(sum(self.item_normal_amount.values()) + sum(self.item_excess_amount.values()))
            self.bill_total_amount_plus_minus = Currency(sum(self.item_plusminus_amount.values()))
            self.bill_plusminus_amount = Currency(self.bill_total_amount_plus_minus * Decimal(percentage)/100)
            self.bill_nettotal_amount = Currency(self.bill_total_amount + self.bill_plusminus_amount)
            if self.prev_bill is not None:
                self.bill_since_prev_amount = Currency(self.bill_nettotal_amount) - CurrencyR(self.prev_bill.bill_nettotal_amount)
            else:
                self.bill_since_prev_amount = Currency(self.bill_nettotal_amount)
            
            self.bill_netpayable_amount = self.bill_since_prev_amount
            for description, amount in self.data.adjustments:
                self.bill_netpayable_amount += Currency(amount)
            self.bill_netpayable_amount = CurrencyR(self.bill_netpayable_amount)
        
        # If bill is a Custom bill
        elif self.data.bill_type == misc.BILL_CUSTOM:
            self.item_qty = self.data.item_qty
            self.item_normal_amount = self.data.item_normal_amount
            self.item_excess_amount = self.data.item_excess_amount
            self.bill_total_amount = Currency(sum(self.data.item_normal_amount.values()) + sum(self.data.item_excess_amount.values()))
            # Compile plusminus items
            for itemno in itemnos:
                if itemno in self.item_normal_amount:
                    item = schedule[itemno]
                    self.item_plusminus_amount[itemno] = Currency(self.item_normal_amount[itemno] if item.percentage else 0)
            self.bill_total_amount_plus_minus = Currency(sum(self.item_plusminus_amount.values()))
            self.bill_plusminus_amount = Currency(self.bill_total_amount * Decimal(percentage)/100)
            self.bill_nettotal_amount = Currency(self.bill_total_amount + self.bill_plusminus_amount)
            
            # For custom bills add adjustments to bill_nettotal_amount
            for description, amount in self.data.adjustments:
                self.bill_nettotal_amount += Currency(amount)
            
            self.bill_since_prev_amount = self.bill_nettotal_amount
            self.bill_netpayable_amount = self.bill_since_prev_amount

    def get_latex_buffer(self, thisbillpath, schedule, project_settings_dict):
        """Return abstract latex buffer"""
        
        percentage_text = project_settings_dict["$cmbtenderpercentage$"]
        percentage = misc.float_from_str(percentage_text)
        
        # Main latex buffer
        latex_buffer = misc.LatexFile()
        
        # Get required datas
        itemnos = schedule.get_itemnos()
                
        # Read preamble and abstract opening into main buffer
        latex_buffer.add_preffix_from_file(misc.abs_path('latex', 'preamble.tex'))
        latex_buffer.add_suffix_from_file(misc.abs_path('latex', 'abstractopening.tex'))

        # Setup substitution dictionary
        bill_local_vars = dict()  # bill substitution dictionary
        bill_local_vars_vanilla = dict()  # bill substitution dictionary
        bill_local_vars['$cmbbookno$'] = self.data.cmb_name
        bill_local_vars['$cmbheading$'] = self.data.title
        bill_local_vars['$cmbtitle$'] = 'ABSTRACT OF COST'
        bill_local_vars['$cmbabstractdate$'] = self.data.bill_date
        bill_local_vars['$cmbstartingpage$'] = str(self.data.starting_page)
        bill_local_vars['$cmbbilltext$'] = self.data.bill_text
        bill_local_vars['$cmbbilltotalamount$'] = str(self.bill_total_amount)
        bill_local_vars['$cmbbillpercentage$'] = percentage_text
        bill_local_vars['$cmbbilltotalamountplusminus$'] = str(self.bill_total_amount_plus_minus)
        bill_local_vars['$cmbbilltotalamountnonplusminus$'] = str(self.bill_total_amount - self.bill_total_amount_plus_minus)
        bill_local_vars['$cmbbillplusminusamount$'] = str(self.bill_plusminus_amount)
        bill_local_vars['$cmbbillnettotalamount$'] = str(self.bill_nettotal_amount)
        if self.prev_bill is not None:
            bill_local_vars['$cmbbillprevamount$'] = str(CurrencyR(self.prev_bill.bill_nettotal_amount))
        else:
            bill_local_vars['$cmbbillprevamount$'] = '0'
        bill_local_vars['$cmbbillsinceprevamount$'] = str(self.bill_since_prev_amount)
        bill_local_vars['$cmbbillsinceprevamountR$'] = str(CurrencyR(self.bill_since_prev_amount))
        bill_local_vars['$cmbbillnetpayableamount$'] = str(CurrencyR(self.bill_netpayable_amount))
        
        if percentage:
            bill_local_vars_vanilla['$billpercentageflag$'] = '\iftrue'
        else:
            bill_local_vars_vanilla['$billpercentageflag$'] = '\iffalse'
        if self.data.bill_type == misc.BILL_FINAL:
            bill_local_vars_vanilla['$finalbillflag$'] = '\iftrue'
        else:
            bill_local_vars_vanilla['$finalbillflag$'] = '\iffalse'
        if self.data.adjustments:
            bill_local_vars_vanilla['$billadjustmentsflag$'] = '\iftrue'
        else:
            bill_local_vars_vanilla['$billadjustmentsflag$'] = '\iffalse'
        bill_local_vars_vanilla['$completionaloneflag$'] = '\iftrue'
            
        # Loop over each item in schedule
        for itemno in itemnos:
            item = schedule[itemno]
            # If item measured, include in bill
            if itemno in self.item_qty and self.item_qty[itemno]:
                # Setup required values
                qty_items = self.item_qty[itemno]
                cmb_refs = self.item_cmb_ref[itemno]
                item_paths = self.item_paths[itemno]
                item_local_vars = {}
                item_local_vars_vanilla = {}
                if self.item_excess_qty[itemno] > 0:
                    excess_flag = '\iftrue'
                else:
                    excess_flag = '\iffalse'
                
                # Include records of each item to bill
                latex_records = misc.LatexFile()
                for qty_item, cmb_ref, item_path in zip(qty_items, cmb_refs, item_paths):
                    if qty_item != 0 or (qty_item == 0 and cmb_ref != -1):
                        # Setup variables
                        item_record_vars = {}
                        item_record_vars['$cmbqtybf$'] = str(round(qty_item,3))
                        item_record_vars['$cmbunit$'] = str(item.unit)
                        if cmb_ref != -1:  # if not prev abstract
                            path_str = str([item_path[0], item_path[1], item_path[2]]) + ':' + str(item_path[3] + 1)
                            item_record_vars['$cmbbf$'] = 'ref:meas:' + path_str
                            item_record_vars['$cmblabel$'] = 'ref:abs:' + path_str
                            item_record_vars['$cmbnormalbillflag$'] = 'iftrue'
                        else:  # if prev abstract
                            if self.prev_bill.data.bill_type in (misc.BILL_NORMAL, misc.BILL_FINAL):
                                hashval = hashlib.sha1(itemno.encode('utf-8')).hexdigest()
                                item_record_vars['$cmbbf$'] = 'ref:abs:abs:' + str([self.data.prev_bill, hashval])
                                item_record_vars['$cmblabel$'] = ''
                                item_record_vars['$cmbnormalbillflag$'] = 'iftrue'
                            elif self.prev_bill.data.bill_type == misc.BILL_CUSTOM:
                                item_record_vars['$cmbbf$'] = self.prev_bill.data.cmb_name
                                item_record_vars['$cmblabel$'] = ''
                                item_record_vars['$cmbnormalbillflag$'] = 'iffalse'

                        # Read in latex records to buffer
                        latex_record = misc.LatexFile()
                        latex_record.add_suffix_from_file(misc.abs_path('latex', 'abstractitemelement.tex'))
                        # Make item substitutions in buffer
                        latex_record.replace_and_clean(item_record_vars)
                        # Add to master buffer
                        latex_records += latex_record

                # Add all values to substitution dict
                item_local_vars['$cmbitemno$'] = str(item.itemno)
                item_local_vars['$cmbdescription$'] = str(item.extended_description_limited)
                item_local_vars['$cmbunit$'] = str(item.unit)
                item_local_vars['$cmbrate$'] = str(item.rate)
                item_local_vars['$cmbexcesspercent$'] = str(item.excess_rate_percent)
                item_local_vars['$cmbtotalqty$'] = str(round(sum(qty_items),3))
                item_local_vars['$cmbnormalqty$'] = str(round(self.item_normal_qty[itemno],3))
                item_local_vars['$cmbexcessqty$'] = str(round(self.item_excess_qty[itemno],3))
                item_local_vars['$cmbexcessrate$'] = str(self.data.item_excess_rates[itemno])
                item_local_vars['$cmbnormalpr$'] = str(
                    Currency(self.data.item_part_percentage[itemno] * 0.01 * schedule[itemno].rate))
                item_local_vars['$cmbexcesspr$'] = str(
                    Currency(self.data.item_excess_part_percentage[itemno] * 0.01 * self.data.item_excess_rates[itemno]))
                item_local_vars['$cmbnormalamount$'] = str(self.item_normal_amount[itemno])
                item_local_vars['$cmbexcessamount$'] = str(self.item_excess_amount[itemno])
                
                hashval = hashlib.sha1(itemno.encode('utf-8')).hexdigest()
                item_local_vars['$cmbabslabel$'] = 'ref:abs:abs:' + str(thisbillpath + [hashval])

                item_local_vars_vanilla['$cmbexcessflag$'] = excess_flag
                item_local_vars_vanilla['$cmbrecords$'] = latex_records.latex_buffer

                # Write entries
                latex_buffer.add_suffix_from_file(misc.abs_path('latex', 'abstractitem.tex')) # Read item template
                # Make item substitutions
                latex_buffer.replace_and_clean(item_local_vars)
                latex_buffer.replace(item_local_vars_vanilla)
                
            # If item not measured, include in bill if final bill
            elif self.data.bill_type == misc.BILL_FINAL:
                item_local_vars = {}
                item_local_vars_vanilla = {}
                excess_flag = '\iffalse'

                # Add all values to substitution dict
                item_local_vars['$cmbitemno$'] = str(item.itemno)
                item_local_vars['$cmbdescription$'] = str(item.extended_description_limited)
                item_local_vars['$cmbunit$'] = str(item.unit)
                item_local_vars['$cmbrate$'] = str(item.rate)
                item_local_vars['$cmbexcesspercent$'] = str(item.excess_rate_percent)
                item_local_vars['$cmbtotalqty$'] = str(round(0,3))
                item_local_vars['$cmbnormalqty$'] = str(round(self.item_normal_qty[itemno],3))
                item_local_vars['$cmbexcessqty$'] = str(round(self.item_excess_qty[itemno],3))
                item_local_vars['$cmbexcessrate$'] = str(self.data.item_excess_rates[itemno])
                item_local_vars['$cmbnormalpr$'] = str(
                    Currency(self.data.item_part_percentage[itemno] * 0.01 * schedule[itemno].rate))
                item_local_vars['$cmbexcesspr$'] = str(
                    Currency(self.data.item_excess_part_percentage[itemno] * 0.01 * self.data.item_excess_rates[itemno]))
                item_local_vars['$cmbnormalamount$'] = str(self.item_normal_amount[itemno])
                item_local_vars['$cmbexcessamount$'] = str(self.item_excess_amount[itemno])
                
                hashval = hashlib.sha1(itemno.encode('utf-8')).hexdigest()
                item_local_vars['$cmbabslabel$'] = 'ref:abs:abs:' + str(thisbillpath + [hashval])

                item_local_vars_vanilla['$cmbexcessflag$'] = excess_flag
                item_local_vars_vanilla['$cmbrecords$'] = ''

                # Write entries
                latex_buffer.add_suffix_from_file(misc.abs_path('latex', 'abstractitem.tex')) # Read item template
                # Make item substitutions
                latex_buffer.replace_and_clean(item_local_vars)
                latex_buffer.replace(item_local_vars_vanilla)
        
        # Add sum blocks
        latex_buffer.add_suffix_from_file(misc.abs_path('latex', 'endabstract1.tex'))
        
        # Add adjustments
        for description, amount in self.data.adjustments:
            item_local_vars = {}            
            # Add all values to substitution dict
            item_local_vars['$cmbbilladjdesc$'] = description
            item_local_vars['$cmbbilladjamount$'] = str(CurrencyR(amount))
            # Write entries
            latex_buffer.add_suffix_from_file(misc.abs_path('latex', 'abstractadjustmentitem.tex')) # Read item template
            # Make item substitutions
            latex_buffer.replace_and_clean(item_local_vars)
                
        # Add abstract end latex block
        latex_buffer.add_suffix_from_file(misc.abs_path('latex', 'endabstract.tex'))
        latex_buffer.replace_and_clean(bill_local_vars)
        latex_buffer.replace(bill_local_vars_vanilla)
        
        return latex_buffer

    def get_latex_buffer_bill(self, schedule, project_settings_dict):
        """Return bill latex buffer"""
        
        percentage_text = project_settings_dict["$cmbtenderpercentage$"]
        percentage = misc.float_from_str(percentage_text)
        if self.prev_bill is not None:
            total_sp = self.bill_total_amount - self.prev_bill.bill_total_amount
            total_plus_minus_sp = self.bill_total_amount_plus_minus - self.prev_bill.bill_total_amount_plus_minus
            sinceprev_calc = self.bill_nettotal_amount - self.prev_bill.bill_nettotal_amount  # Since previous amount without rounding
            plusminus_sp = self.bill_plusminus_amount - self.prev_bill.bill_plusminus_amount
        else:
            total_sp = self.bill_total_amount
            total_plus_minus_sp = 0
            sinceprev_calc = self.bill_nettotal_amount
            plusminus_sp = self.bill_plusminus_amount
        
        # Main latex buffer
        latex_buffer = misc.LatexFile()
        
        # Get required datas
        itemnos = schedule.get_itemnos()
        
        # Read prefix
        latex_buffer.add_suffix_from_file(misc.abs_path('latex','billopening.tex'))
        
        # Setup substitution dictionary
        bill_local_vars = dict()  # bill substitution dictionary
        bill_local_vars_vanilla = dict()  # bill substitution dictionary
        bill_local_vars['$cmbbookno$'] = self.data.cmb_name
        bill_local_vars['$cmbheading$'] = self.data.title
        bill_local_vars['$cmbabstractdate$'] = self.data.bill_date
        
        if self.prev_bill:
            bill_local_vars['$cmbheadingold$'] = self.prev_bill.data.title
        else:
            bill_local_vars['$cmbheadingold$'] = ''
        
        bill_local_vars['$cmbbilltotalamount$'] = str(self.bill_total_amount)
        bill_local_vars['$cmbbillpercentage$'] = percentage_text
        bill_local_vars['$cmbbilltotalamountplusminus$'] = str(self.bill_total_amount_plus_minus)
        bill_local_vars['$cmbbilltotalamountnonplusminus$'] = str(self.bill_total_amount - self.bill_total_amount_plus_minus)
        bill_local_vars['$cmbbillplusminusamount$'] = str(self.bill_plusminus_amount)
        bill_local_vars['$cmbbillnettotalamount$'] = str(self.bill_nettotal_amount)
        if self.prev_bill is not None:
            bill_local_vars['$cmbbillprevamount$'] = str(CurrencyR(self.prev_bill.bill_nettotal_amount))
        else:
            bill_local_vars['$cmbbillprevamount$'] = '0'
        bill_local_vars['$cmbbillsinceprevamount$'] = str(self.bill_since_prev_amount)
        bill_local_vars['$cmbbillsinceprevamountR$'] = str(CurrencyR(self.bill_since_prev_amount))
        bill_local_vars['$cmbbilltotalamountsp$'] = str(total_sp)
        bill_local_vars['$cmbbilltotalamountplusminussp$'] = str(total_plus_minus_sp)
        bill_local_vars['$cmbbilltotalamountnonplusminussp$'] = str(total_sp-total_plus_minus_sp)
        bill_local_vars['$cmbbillplusminusamountsp$'] = str(plusminus_sp)
        bill_local_vars['$cmbbillsinceprevamountcalc$'] = str(sinceprev_calc)
        bill_local_vars['$cmbbillnetpayableamount$'] = str(CurrencyR(self.bill_netpayable_amount))
        
        if percentage:
            bill_local_vars_vanilla['$billpercentageflag$'] = '\iftrue'
        else:
            bill_local_vars_vanilla['$billpercentageflag$'] = '\iffalse'
        if self.data.adjustments:
            bill_local_vars_vanilla['$billadjustmentsflag$'] = '\iftrue'
        else:
            bill_local_vars_vanilla['$billadjustmentsflag$'] = '\iffalse'
        if self.data.bill_type == misc.BILL_FINAL:
            bill_local_vars_vanilla['$finalbillflag$'] = '\iftrue'
        else:
            bill_local_vars_vanilla['$finalbillflag$'] = '\iffalse'
        
        # Write each item to bill
        for itemno in itemnos:
            item = schedule[itemno]
            
            # If item measured, include in bill
            if itemno in self.item_qty and self.item_qty[itemno]:
                # Setup required values
                qty_items = self.item_qty[itemno]
                item_paths = self.item_paths[itemno]
                
                if self.prev_bill is not None and itemno in self.prev_bill.item_normal_amount:
                    sprev_item_normal_amount = Currency(self.item_normal_amount[itemno] \
                                               - Decimal(self.prev_bill.item_normal_amount[itemno]))
                    sprev_item_excess_amount = Currency(self.item_excess_amount[itemno] \
                                               - Decimal(self.prev_bill.item_excess_amount[itemno]))
                else:
                    sprev_item_normal_amount = self.item_normal_amount[itemno]
                    sprev_item_excess_amount = self.item_excess_amount[itemno]
                    
                item_local_vars = {}
                item_local_vars_vanilla = {}

                if self.item_excess_qty[itemno] > 0:
                    excess_flag = '\iftrue'
                else:
                    excess_flag = '\iffalse'

                # Add all values to substitution dict
                item_local_vars['$cmbitemno$'] = str(item.itemno)
                item_local_vars['$cmbdescription$'] = str(item.extended_description_limited)
                item_local_vars['$cmbunit$'] = str(item.unit)
                item_local_vars['$cmbrate$'] = str(item.rate)
                item_local_vars['$cmbexcesspercent$'] = str(item.excess_rate_percent)
                item_local_vars['$cmbtotalqty$'] = str(round(sum(qty_items),3))
                item_local_vars['$cmbnormalqty$'] = str(round(self.item_normal_qty[itemno],3))
                item_local_vars['$cmbexcessqty$'] = str(round(self.item_excess_qty[itemno],3))
                item_local_vars['$cmbexcessrate$'] = str(self.data.item_excess_rates[itemno])
                item_local_vars['$cmbnormalpr$'] = str(
                    Currency(self.data.item_part_percentage[itemno] * 0.01 * schedule[itemno].rate))
                item_local_vars['$cmbexcesspr$'] = str(
                    Currency(self.data.item_excess_part_percentage[itemno] * 0.01 * self.data.item_excess_rates[itemno]))
                item_local_vars['$cmbnormalamount$'] = str(self.item_normal_amount[itemno])
                item_local_vars['$cmbexcessamount$'] = str(self.item_excess_amount[itemno])
                item_local_vars['$cmbnormalsinceprevamount$'] = str(sprev_item_normal_amount)
                item_local_vars['$cmbexcesssinceprevamount$'] = str(sprev_item_excess_amount)

                item_local_vars_vanilla['$cmbexcessflag$'] = excess_flag
                
                # Write entries
                latex_buffer.add_suffix_from_file(misc.abs_path('latex', 'billitem.tex')) # read item template
                # Make item substitutions
                latex_buffer.replace_and_clean(item_local_vars)
                latex_buffer.replace(item_local_vars_vanilla)
            
            # If item not measured, include in bill if final bill
            elif self.data.bill_type == misc.BILL_FINAL:
                # Setup required values
                sprev_item_normal_amount = self.item_normal_amount[itemno]
                sprev_item_excess_amount = self.item_excess_amount[itemno]
                item_local_vars = {}
                item_local_vars_vanilla = {}
                excess_flag = '\iffalse'

                # Add all values to substitution dict
                item_local_vars['$cmbitemno$'] = str(item.itemno)
                item_local_vars['$cmbdescription$'] = str(item.extended_description_limited)
                item_local_vars['$cmbunit$'] = str(item.unit)
                item_local_vars['$cmbrate$'] = str(item.rate)
                item_local_vars['$cmbexcesspercent$'] = str(item.excess_rate_percent)
                item_local_vars['$cmbtotalqty$'] = str(round(0,3))
                item_local_vars['$cmbnormalqty$'] = str(round(self.item_normal_qty[itemno],3))
                item_local_vars['$cmbexcessqty$'] = str(round(self.item_excess_qty[itemno],3))
                item_local_vars['$cmbexcessrate$'] = str(self.data.item_excess_rates[itemno])
                item_local_vars['$cmbnormalpr$'] = str(
                    Currency(self.data.item_part_percentage[itemno] * 0.01 * schedule[itemno].rate))
                item_local_vars['$cmbexcesspr$'] = str(
                    Currency(self.data.item_excess_part_percentage[itemno] * 0.01 * self.data.item_excess_rates[itemno]))
                item_local_vars['$cmbnormalamount$'] = str(self.item_normal_amount[itemno])
                item_local_vars['$cmbexcessamount$'] = str(self.item_excess_amount[itemno])
                item_local_vars['$cmbnormalsinceprevamount$'] = str(sprev_item_normal_amount)
                item_local_vars['$cmbexcesssinceprevamount$'] = str(sprev_item_excess_amount)

                item_local_vars_vanilla['$cmbexcessflag$'] = excess_flag
                
                # Write entries
                latex_buffer.add_suffix_from_file(misc.abs_path('latex', 'billitem.tex')) # read item template
                # Make item substitutions
                latex_buffer.replace_and_clean(item_local_vars)
                latex_buffer.replace(item_local_vars_vanilla)
        
        # Add sum blocks
        latex_buffer.add_suffix_from_file(misc.abs_path('latex', 'endbill1.tex'))
        
        # Add adjustments
        for description, amount in self.data.adjustments:
            item_local_vars = {}            
            # Add all values to substitution dict
            item_local_vars['$cmbbilladjdesc$'] = description
            item_local_vars['$cmbbilladjamount$'] = str(CurrencyR(amount))
            # Write entries
            latex_buffer.add_suffix_from_file(misc.abs_path('latex', 'billadjustmentitem.tex')) # Read item template
            # Make item substitutions
            latex_buffer.replace_and_clean(item_local_vars)
                
        # Add bill end latex block
        latex_buffer.add_suffix_from_file(misc.abs_path('latex', 'endbill.tex'))
        
        # Make replacements
        latex_buffer.replace_and_clean(bill_local_vars)
        latex_buffer.replace(bill_local_vars_vanilla)
        
        return latex_buffer

    def export_spreadsheet_bill(self, filename, project_settings_dict, schedule, cmbs):
        """Export bill to spreadsheet file"""
        spreadsheet = Workbook()
        
        estimated_cost = misc.float_from_str(project_settings_dict["$cmbestimatedcost$"])
        percentage = misc.float_from_str(project_settings_dict["$cmbtenderpercentage$"])

        ## SHEET 1 DATA
        sheet = spreadsheet.active
        sheet.title = 'Data'
        sheet_head = ['Agmnt.No','Description','Unit','Total Qty','Below Dev Qty',
                      'Above Dev Qty','Agmnt FR','Agmnt PR','Excess FR','Excess PR',
                      'Total Bel Dev','Total Above Dev',
                      'Since Prev below Dev','Since Prev Above Dev','Dev Limit']
        for c,head in enumerate(sheet_head):
            sheet.cell(row=1,column=c+1).value = head
            sheet.cell(row=1,column=c+1).font = Font(bold=True)
            sheet.cell(row=1,column=c+1).alignment = Alignment(wrap_text=True, vertical='center')
        for count in range(schedule.length()):
            item = schedule.get_item_by_index(count)
            
            sheet.cell(row=count+2, column=1).value = item.itemno
            sheet.cell(row=count+2, column=2).value = item.description
            sheet.cell(row=count+2, column=2).alignment = Alignment(wrap_text=True)
            sheet.cell(row=count+2, column=3).value = item.unit
            sheet.cell(row=count+2, column=7).value = item.rate
            sheet.cell(row=count+2, column=15).value = item.excess_rate_percent
            
            # Fill in values for measured items
            if item.itemno in self.item_qty:
                itemno = item.itemno
                sheet.cell(row=count+2, column=4).value = sum(self.item_qty[itemno])
                sheet.cell(row=count+2, column=5).value = self.item_normal_qty[itemno]
                sheet.cell(row=count+2, column=6).value = self.item_excess_qty[itemno]
                sheet.cell(row=count+2, column=8).value = Currency(self.data.item_part_percentage[itemno] *
                                                 0.01 * schedule[itemno].rate)
                sheet.cell(row=count+2, column=9).value = self.data.item_excess_rates[itemno]
                sheet.cell(row=count+2, column=10).value = Currency(self.data.item_excess_part_percentage[itemno] *
                                                 0.01 * self.data.item_excess_rates[itemno])
                sheet.cell(row=count+2, column=11).value = self.item_normal_amount[itemno]
                sheet.cell(row=count+2, column=12).value = self.item_excess_amount[itemno]
                if self.prev_bill is not None and item.itemno in self.prev_bill.item_normal_amount:
                    sprev_item_normal_amount = self.item_normal_amount[itemno] \
                                               - Decimal(self.prev_bill.item_normal_amount[itemno])
                    sprev_item_excess_amount = self.item_excess_amount[itemno] \
                                               - Decimal(self.prev_bill.item_excess_amount[itemno])
                else:
                    sprev_item_normal_amount = self.item_normal_amount[itemno]
                    sprev_item_excess_amount = self.item_excess_amount[itemno]
                sheet.cell(row=count+2, column=13).value = sprev_item_normal_amount
                sheet.cell(row=count+2, column=14).value = sprev_item_excess_amount

        # Sheet formatings
        sheet.column_dimensions['B'].width = 50
        sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
        sheet.page_setup.paperSize = sheet.PAPERSIZE_LEGAL
        sheet.page_setup.fitToHeight = 99
        sheet.page_setup.fitToWidth = 1


        ## SHEET 2 DEVIATION STATEMENT
        sheet2 = spreadsheet.create_sheet()
        sheet2.title = 'Deviation'
        template = load_workbook(filename = misc.abs_path('ods_templates','dev.xlsx'))
        rowend = 12
        colend = 18
        rowend_end = 3
        
        # Copy all from dev start
        template_start_sheet = template['start']
        for row in range(1,rowend+1):
            for column in range(1,colend+1):
                sheet2.cell(row=row, column=column).value = template_start_sheet.cell(row=row, column=column).value
                sheet2.cell(row=row, column=column).font = copy.copy(template_start_sheet.cell(row=row, column=column).font)
                sheet2.cell(row=row, column=column).border = copy.copy(template_start_sheet.cell(row=row, column=column).border)
                sheet2.cell(row=row, column=column).alignment = copy.copy(template_start_sheet.cell(row=row, column=column).alignment)
        
        # Copy all values
        for count in range(schedule.length()):
            item = schedule.get_item_by_index(count)
            sheet2.cell(row=count+rowend+1, column=1).value = '=IF(INDIRECT(ADDRESS(ROW(),5))<>0,MAX(INDIRECT("A$12:" & ADDRESS(ROW()-1,1)))+1,"")'
            sheet2.cell(row=count+rowend+1, column=2).value = item.itemno
            sheet2.cell(row=count+rowend+1, column=3).value = item.description
            if item.qty > 0:
                sheet2.cell(row=count+rowend+1, column=4).value = item.unit
                sheet2.cell(row=count+rowend+1, column=5).value = item.qty
                # Fill in values for measured items
                if item.itemno in self.item_qty:
                    itemno = item.itemno
                    percent_dev = round((sum(self.item_qty[itemno]) - item.qty)/item.qty*100,2) if item.qty != 0 else 0
                    sheet2.cell(row=count+rowend+1, column=6).value = sum(self.item_qty[itemno])
                    sheet2.cell(row=count+rowend+1, column=7).value = sum(self.item_qty[itemno]) - item.qty
                    if self.item_normal_qty[itemno]-item.qty > 0:
                        sheet2.cell(row=count+rowend+1, column=9).value = self.item_normal_qty[itemno] - item.qty
                    if self.item_excess_qty[itemno] > 0:
                        sheet2.cell(row=count+rowend+1, column=10).value = self.item_excess_qty[itemno]
                    sheet2.cell(row=count+rowend+1, column=11).value = \
                            '=IF(ABS(INDIRECT(ADDRESS(ROW(),8)))>' + str(misc.DEV_LIMIT_STATEMENT) + ',' + \
                            str(self.item_normal_qty[itemno] - item.qty) + ',0)'
                    sheet2.cell(row=count+rowend+1, column=15).value = self.data.item_excess_rates[itemno]
                # Fill in values for items not measured
                else:
                    percent_dev = -100
                    sheet2.cell(row=count+rowend+1, column=6).value = 0
                    sheet2.cell(row=count+rowend+1, column=7).value = -item.qty
                    sheet2.cell(row=count+rowend+1, column=9).value = -item.qty
                    sheet2.cell(row=count+rowend+1, column=10).value = 0
                    sheet2.cell(row=count+rowend+1, column=11).value = \
                            '=IF(ABS(INDIRECT(ADDRESS(ROW(),8)))>' + str(misc.DEV_LIMIT_STATEMENT) + ',' + \
                            str(self.item_normal_qty[itemno] - item.qty) + ',0)'
                    sheet2.cell(row=count+rowend+1, column=15).value = self.data.item_excess_rates[itemno]
                if item.percentage:
                    sheet2.cell(row=count+rowend+1, column=12).value = Currency(item.rate * (100 + percentage)/100)
                else:
                    sheet2.cell(row=count+rowend+1, column=12).value = item.rate
                sheet2.cell(row=count+rowend+1, column=8).value = percent_dev
                # Fill in formulas
                sheet2.cell(row=count+rowend+1, column=13).value = \
                    '=ROUND(INDIRECT(ADDRESS(ROW(),11))*INDIRECT(ADDRESS(ROW(),12)), 2)'
                sheet2.cell(row=count+rowend+1, column=14).value = '=INDIRECT(ADDRESS(ROW(),10))'
                sheet2.cell(row=count+rowend+1, column=16).value = \
                    '=ROUND(INDIRECT(ADDRESS(ROW(),14))*INDIRECT(ADDRESS(ROW(),15)), 2)'
                sheet2.cell(row=count+rowend+1, column=17).value = \
                    '=ABS(INDIRECT(ADDRESS(ROW(),13)))+ABS(INDIRECT(ADDRESS(ROW(),16)))'
            # Formatings
            sheet2.cell(row=count+rowend+1, column=1).alignment = Alignment(horizontal='center')
            sheet2.cell(row=count+rowend+1, column=2).alignment = Alignment(horizontal='center')
            sheet2.cell(row=count+rowend+1, column=3).alignment = Alignment(wrap_text=True)
            
        # Copy all from dev end
        template_end_sheet = template['end']
        for row,row_ in zip(list(range(rowend+schedule.length()+1,rowend+schedule.length()+rowend_end+1)),list(range(1,rowend_end+1))):
            for column,column_ in zip(list(range(1,colend+1)),list(range(1,colend+1))):
                sheet2.cell(row=row, column=column).value = template_end_sheet.cell(row=row_, column=column_).value
                sheet2.cell(row=row, column=column).font = copy.copy(template_end_sheet.cell(row=row, column=column).font)
                sheet2.cell(row=row, column=column).border = copy.copy(template_end_sheet.cell(row=row, column=column).border)
                sheet2.cell(row=row, column=column).alignment = copy.copy(template_end_sheet.cell(row=row, column=column).alignment)
        sheet2.cell(row=rowend+schedule.length()+1, column=colend-1).value = \
            '=SUM('+get_column_letter(colend-1)+str(rowend+1)+':'+get_column_letter(colend-1)+\
            str(rowend+schedule.length())+')'
            
        # sheet2 formatings
        for column in range(1,colend+1):
            # copy coumn widths
            sheet2.column_dimensions[get_column_letter(column)].width = \
                template_start_sheet.column_dimensions[get_column_letter(column)].width
        sheet2.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
        sheet2.page_setup.paperSize = sheet.PAPERSIZE_LEGAL
        sheet2.page_setup.fitToHeight = 99
        sheet2.page_setup.fitToWidth = 1
        sheet2.print_options.horizontalCentered = True

        # Copy all data
        sheet2['C3'] = project_settings_dict["$cmbnameofwork$"]
        sheet2['C4'] = project_settings_dict["$cmbagency$"]
        sheet2['C5'] = project_settings_dict["$cmbagmntno$"]
        sheet2['C6'] = Currency(estimated_cost * (percentage + 100)/100)
        
        
        ## SHEET 3 ABSTRACT
        sheet = spreadsheet.create_sheet()
        sheet.title = 'Abs'
        template = load_workbook(filename = misc.abs_path('ods_templates','abs.xlsx'))
        rowend = 12
        colend = 7
        rowend_end = 8
        
        # Copy all from Abstract start
        template_start_sheet = template['start']
        for row in range(1,rowend+1):
            for column in range(1,colend+1):
                sheet.cell(row=row, column=column).value = template_start_sheet.cell(row=row, column=column).value
                sheet.cell(row=row, column=column).font = copy.copy(template_start_sheet.cell(row=row, column=column).font)
                sheet.cell(row=row, column=column).border = copy.copy(template_start_sheet.cell(row=row, column=column).border)
                sheet.cell(row=row, column=column).alignment = copy.copy(template_start_sheet.cell(row=row, column=column).alignment)
                
        # Copy all data
        sheet['C3'] = 'ABSTRACT OF COST FOR ' + self.data.title
        sheet['C5'] = project_settings_dict["$cmbnameofwork$"]
        sheet['C6'] = project_settings_dict["$cmbsituation$"]
        sheet['C7'] = project_settings_dict["$cmbagmntno$"]
        sheet['C8'] = project_settings_dict["$cmbagency$"]
        sheet['C9'] = project_settings_dict["$cmbdateofstart$"]
        sheet['C10'] = project_settings_dict["$cmbdateofstartasperagmnt$"]
        sheet['C11'] = ''
        sheet['C12'] = self.data.bill_date
        
        # Copy abstract items
        
        itemnos = schedule.get_itemnos()
        row_item = rowend+2
        rownums_per_items = []
        rownums_nonper_items = []
        for itemno in itemnos:
            item = schedule[itemno]
            # If item measured, include in bill
            if itemno in self.item_qty and self.item_qty[itemno]:
                # Setup required values
                qty_items = self.item_qty[itemno]
                cmb_refs = self.item_cmb_ref[itemno]
                item_paths = self.item_paths[itemno]
                
                if self.prev_bill is not None and itemno in self.prev_bill.item_normal_amount:
                    sprev_item_normal_amount = self.item_normal_amount[itemno] \
                                               - Decimal(self.prev_bill.item_normal_amount[itemno])
                    sprev_item_excess_amount = self.item_excess_amount[itemno] \
                                               - Decimal(self.prev_bill.item_excess_amount[itemno])
                else:
                    sprev_item_normal_amount = self.item_normal_amount[itemno]
                    sprev_item_excess_amount = self.item_excess_amount[itemno]
                    
                # Copy data
                sheet['A' + str(row_item)] = itemno
                sheet['B' + str(row_item)] = item.extended_description
                sheet['A' + str(row_item)].alignment = Alignment(horizontal='center')
                sheet['B' + str(row_item)].alignment = Alignment(wrap_text=True)
                row_item += 1
                
                # Include records of each item to bill
                if qty_items:
                    for qty_item, cmb_ref, item_path in zip(qty_items, cmb_refs, item_paths):
                        # Print zero quantity items only for final bill
                        if qty_item != 0 or (qty_item == 0 and cmb_ref != -1):
                            if cmb_ref != -1:  # if not prev abstract
                                sheet['B' + str(row_item)] = 'Qty B/F Mb.No.' + str(cmbs[cmb_ref].name) + ' Pg.No.'
                                sheet['C' + str(row_item)] = qty_item
                                sheet['D' + str(row_item)] = item.unit
                            else:  # if prev abstract
                                sheet['B' + str(row_item)] = 'Qty B/F Prev Abs Mb.No.' + self.prev_bill.data.cmb_name + ' Pg.No.'
                                sheet['C' + str(row_item)] = qty_item
                                sheet['D' + str(row_item)] = item.unit
                            row_item += 1
                        
                if self.item_excess_qty[itemno] > 0:
                    sheet['B' + str(row_item)] = 'TOTAL'
                    sheet['C' + str(row_item)] = sum(qty_items)
                    sheet['D' + str(row_item)] = item.unit
                    row_item += 1
                    sheet['B' + str(row_item)] = 'Qty upto deviation limit of ' + str(item.excess_rate_percent) + '%'
                    sheet['C' + str(row_item)] = self.item_normal_qty[itemno]
                    sheet['D' + str(row_item)] = item.unit
                    sheet['E' + str(row_item)] = item.rate
                    sheet['F' + str(row_item)] = Currency(item.rate*self.data.item_part_percentage[itemno]/100)
                    sheet['G' + str(row_item)] = '=ROUND(C'+ str(row_item) + '*F' + str(row_item) + ',2)'
                    if item.percentage:
                        rownums_per_items.append(row_item)
                    else:
                        rownums_nonper_items.append(row_item)
                    row_item += 1
                    sheet['B' + str(row_item)] = 'Qty above deviation limit of ' + str(item.excess_rate_percent) + '%'
                    sheet['C' + str(row_item)] = self.item_excess_qty[itemno]
                    sheet['D' + str(row_item)] = item.unit
                    sheet['E' + str(row_item)] = self.data.item_excess_rates[itemno]
                    sheet['F' + str(row_item)] = Currency(self.data.item_excess_rates[itemno]*self.data.item_excess_part_percentage[itemno]/100)
                    sheet['G' + str(row_item)] = '=ROUND(C'+ str(row_item) + '*F' + str(row_item) + ',2)'
                    rownums_nonper_items.append(row_item)
                    row_item += 2
                else:
                    sheet['B' + str(row_item)] = 'TOTAL'
                    sheet['C' + str(row_item)] = sum(self.item_qty[itemno])
                    sheet['D' + str(row_item)] = item.unit
                    sheet['E' + str(row_item)] = item.rate
                    sheet['F' + str(row_item)] = Currency(item.rate*self.data.item_part_percentage[itemno]/100)
                    sheet['G' + str(row_item)] = '=ROUND(C'+ str(row_item) + '*F' + str(row_item) + ',2)'
                    if item.percentage:
                        rownums_per_items.append(row_item)
                    else:
                        rownums_nonper_items.append(row_item)
                    row_item += 2
            
            # If item not measured, include in bill if final bill
            elif self.data.bill_type == misc.BILL_FINAL:
                # Copy data
                sheet['A' + str(row_item)] = itemno
                sheet['B' + str(row_item)] = item.extended_description
                sheet['A' + str(row_item)].alignment = Alignment(horizontal='center')
                sheet['B' + str(row_item)].alignment = Alignment(wrap_text=True)
                row_item += 1
                sheet['B' + str(row_item)] = 'Item not executed'
                sheet['C' + str(row_item)] = 0
                sheet['D' + str(row_item)] = item.unit
                sheet['E' + str(row_item)] = item.rate
                sheet['F' + str(row_item)] = Currency(item.rate*self.data.item_part_percentage[itemno]/100)
                sheet['G' + str(row_item)] = '=ROUND(C'+ str(row_item) + '*F' + str(row_item) + ',2)'
                if item.percentage:
                    rownums_per_items.append(row_item)
                else:
                    rownums_nonper_items.append(row_item)
                row_item += 1
            
        # Copy all from abs end
        template_end_sheet = template['end']
        for row in range(1,rowend_end+1):
            for column in range(1,colend+1):
                sheet.cell(row=row+row_item, column=column).value = template_end_sheet.cell(row=row, column=column).value
                sheet.cell(row=row, column=column).font = copy.copy(template_end_sheet.cell(row=row, column=column).font)
                sheet.cell(row=row, column=column).border = copy.copy(template_end_sheet.cell(row=row, column=column).border)
                sheet.cell(row=row, column=column).alignment = copy.copy(template_end_sheet.cell(row=row, column=column).alignment)
        
        # Fill in values
        if rownums_per_items:
            percentage_am_eq = '=ROUND((' + '+'.join(['G' + str(x) for x in rownums_per_items]) + '), 2)'
        else:
            percentage_am_eq = 0
        if rownums_nonper_items:
            nonpercentage_am_eq = '=ROUND((' + '+'.join(['G' + str(x) for x in rownums_nonper_items]) + '), 2)'
        else:
            nonpercentage_am_eq = 0
        sheet.cell(row=1+row_item, column=7).value = percentage_am_eq
        sheet.cell(row=2+row_item, column=7).value = nonpercentage_am_eq
        sheet.cell(row=3+row_item, column=7).value = '=G' + str(row_item+1) + '+G' + str(row_item+2)  # Up to date amount
        sheet.cell(row=4+row_item, column=7).value = '=ROUND(G' + str(row_item+1) + '*' + str(percentage/100) + ', 2)'  # Percentage amount
        sheet.cell(row=5+row_item, column=7).value = '=G' + str(row_item+3) + '+G' + str(row_item+4)  # Net up to date amount
        if self.prev_bill != None:
            sheet.cell(row=6+row_item, column=7).value = CurrencyR(self.prev_bill.bill_nettotal_amount)  # Previous bill amount
        sheet.cell(row=7+row_item, column=7).value = '=ROUND(G' + str(row_item+5) + '-G' + str(row_item+6) + ', 2)'  # Since previous amount
        sheet.cell(row=8+row_item, column=7).value = '=ROUND(G' + str(row_item+7) + ', 0)'  # Since previous amount rounded
        
        sheet.cell(row=4+row_item, column=2).value = sheet.cell(row=4+row_item, column=2).value + ' @ ' + str(percentage) + '%'  # Add percentage value
        row_item_sp = row_item + 8
        row_item += 10
        
        # Add details of Adjustments
        for description, amount in self.data.adjustments:
            if amount != 0:
                sheet['B' + str(row_item)] = description
                sheet['G' + str(row_item)] = Currency(amount)
                row_item += 1
        # Add net amount
        if self.data.adjustments:
            sheet['B' + str(row_item)] = 'NET AMOUNT PAYABLE'
            sheet['G' + str(row_item)] = '=ROUND(SUM(G' + str(row_item_sp) + ':G' + str(row_item-1) + '), 0)'
        
        # Fill in text
        sheet.cell(row=2+row_item, column=2).value = self.data.bill_text
        
        # Abstract formatting
        for column in range(1,colend+1):
            # copy column widths
            sheet.column_dimensions[get_column_letter(column)].width = \
                template_start_sheet.column_dimensions[get_column_letter(column)].width
        sheet.page_setup.orientation = sheet.ORIENTATION_PORTRAIT
        sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
        sheet.page_setup.fitToHeight = 99
        sheet.page_setup.fitToWidth = 1
        sheet.print_options.horizontalCentered = True
        
        
        ## SHEET 4 BILL
        sheet = spreadsheet.create_sheet()
        sheet.title = 'Bill'
        template = load_workbook(filename = misc.abs_path('ods_templates','bill.xlsx'))
        rowend = 7
        colend = 8
        rowend_end = 8
        
        # Copy all from bill start
        template_start_sheet = template['start']
        for row in range(1,rowend+1):
            for column in range(1,colend+1):
                sheet.cell(row=row, column=column).value = template_start_sheet.cell(row=row, column=column).value
                sheet.cell(row=row, column=column).font = copy.copy(template_start_sheet.cell(row=row, column=column).font)
                sheet.cell(row=row, column=column).border = copy.copy(template_start_sheet.cell(row=row, column=column).border)
                sheet.cell(row=row, column=column).alignment = copy.copy(template_start_sheet.cell(row=row, column=column).alignment)
                
        # Copy all data
        sheet['C3'] = 'SCHEDULE OF RATES FOR ' + self.data.title
        sheet['C5'] = project_settings_dict["$cmbnameofwork$"]
        sheet['C6'] = project_settings_dict["$cmbagmntno$"]
        sheet['C7'] = project_settings_dict["$cmbagency$"]
        
        # Copy bill items
        
        itemnos = schedule.get_itemnos()
        rownums_per_items = []
        rownums_nonper_items = []
        row_item = rowend+2
        for itemno in itemnos:
            item = schedule[itemno]
            # If item measured, include in bill
            if itemno in self.item_qty and self.item_qty[itemno]:
                # Setup required values
                qty_items = self.item_qty[itemno]
                cmb_refs = self.item_cmb_ref[itemno]
                item_paths = self.item_paths[itemno]
                
                if self.prev_bill is not None and itemno in self.prev_bill.item_normal_amount:
                    sprev_item_normal_amount = self.item_normal_amount[itemno] \
                                               - Decimal(self.prev_bill.item_normal_amount[itemno])
                    sprev_item_excess_amount = self.item_excess_amount[itemno] \
                                               - Decimal(self.prev_bill.item_excess_amount[itemno])
                else:
                    sprev_item_normal_amount = self.item_normal_amount[itemno]
                    sprev_item_excess_amount = self.item_excess_amount[itemno]
                    
                # Copy data
                sheet['A' + str(row_item)] = itemno
                sheet['B' + str(row_item)] = item.extended_description
                sheet['A' + str(row_item)].alignment = Alignment(horizontal='center')
                sheet['B' + str(row_item)].alignment = Alignment(wrap_text=True)
                row_item += 1
                
                if self.item_excess_qty[itemno] > 0:
                    sheet['B' + str(row_item)] = 'TOTAL'
                    sheet['C' + str(row_item)] = sum(qty_items)
                    sheet['D' + str(row_item)] = item.unit
                    row_item += 1
                    sheet['B' + str(row_item)] = 'Qty upto deviation limit of ' + str(item.excess_rate_percent) + '%'
                    sheet['C' + str(row_item)] = self.item_normal_qty[itemno]
                    sheet['D' + str(row_item)] = item.unit
                    sheet['E' + str(row_item)] = item.rate
                    sheet['F' + str(row_item)] = Currency(item.rate*self.data.item_part_percentage[itemno]/100)
                    sheet['G' + str(row_item)] = '=ROUND(C'+ str(row_item) + '*F' + str(row_item) + ',2)'
                    if self.prev_bill != None and itemno in self.prev_bill.item_normal_amount:
                        sheet['H' + str(row_item)] = self.item_normal_amount[itemno] - Decimal(self.prev_bill.item_normal_amount[itemno])
                    else:
                        sheet['H' + str(row_item)] = self.item_normal_amount[itemno]
                    if item.percentage:
                        rownums_per_items.append(row_item)
                    else:
                        rownums_nonper_items.append(row_item)
                    row_item += 1
                    sheet['B' + str(row_item)] = 'Qty above deviation limit of ' + str(item.excess_rate_percent) + '%'
                    sheet['C' + str(row_item)] = self.item_excess_qty[itemno]
                    sheet['D' + str(row_item)] = item.unit
                    sheet['E' + str(row_item)] = self.data.item_excess_rates[itemno]
                    sheet['F' + str(row_item)] = Currency(self.data.item_excess_rates[itemno]*self.data.item_excess_part_percentage[itemno]/100)
                    sheet['G' + str(row_item)] = '=ROUND(C'+ str(row_item) + '*F' + str(row_item) + ',2)'
                    if self.prev_bill != None and itemno in self.prev_bill.item_excess_amount:
                        sheet['H' + str(row_item)] = self.item_excess_amount[itemno] - Decimal(self.prev_bill.item_excess_amount[itemno])
                    else:
                        sheet['H' + str(row_item)] = self.item_excess_amount[itemno]
                    rownums_nonper_items.append(row_item)
                    row_item += 2
                else:
                    sheet['B' + str(row_item)] = 'TOTAL'
                    sheet['C' + str(row_item)] = sum(self.item_qty[itemno])
                    sheet['D' + str(row_item)] = item.unit
                    sheet['E' + str(row_item)] = item.rate
                    sheet['F' + str(row_item)] = Currency(item.rate*self.data.item_part_percentage[itemno]/100)
                    sheet['G' + str(row_item)] = '=ROUND(C'+ str(row_item) + '*F' + str(row_item) + ',2)'
                    if self.prev_bill != None  and itemno in self.prev_bill.item_normal_amount:
                        sheet['H' + str(row_item)] = self.item_normal_amount[itemno] - Decimal(self.prev_bill.item_normal_amount[itemno])
                    else:
                        sheet['H' + str(row_item)] = self.item_normal_amount[itemno]
                    if item.percentage:
                        rownums_per_items.append(row_item)
                    else:
                        rownums_nonper_items.append(row_item)
                    row_item += 2
            
            # If item not measured, include in bill if final bill
            elif self.data.bill_type == misc.BILL_FINAL:
                # Copy data
                sheet['A' + str(row_item)] = itemno
                sheet['B' + str(row_item)] = item.extended_description
                sheet['A' + str(row_item)].alignment = Alignment(horizontal='center')
                sheet['B' + str(row_item)].alignment = Alignment(wrap_text=True)
                row_item += 1
                sheet['B' + str(row_item)] = 'TOTAL'
                sheet['C' + str(row_item)] = 0
                sheet['D' + str(row_item)] = item.unit
                sheet['E' + str(row_item)] = item.rate
                sheet['F' + str(row_item)] = Currency(item.rate*self.data.item_part_percentage[itemno]/100)
                sheet['G' + str(row_item)] = '=ROUND(C'+ str(row_item) + '*F' + str(row_item) + ',2)'
                sheet['H' + str(row_item)] = 0
                if item.percentage:
                    rownums_per_items.append(row_item)
                else:
                    rownums_nonper_items.append(row_item)
                row_item += 1
                
        # Copy all from bill end
        template_end_sheet = template['end']
        for row in range(1,rowend_end+1):
            for column in range(1,colend+1):
                sheet.cell(row=row+row_item, column=column).value = template_end_sheet.cell(row=row, column=column).value
                sheet.cell(row=row, column=column).font = copy.copy(template_end_sheet.cell(row=row, column=column).font)
                sheet.cell(row=row, column=column).border = copy.copy(template_end_sheet.cell(row=row, column=column).border)
                sheet.cell(row=row, column=column).alignment = copy.copy(template_end_sheet.cell(row=row, column=column).alignment)
        
        # Fill in values
        if rownums_per_items:
            percentage_am_eq = '=ROUND((' + '+'.join(['G' + str(x) for x in rownums_per_items]) + '), 2)'
            percentage_am_eq_sp = '=ROUND((' + '+'.join(['H' + str(x) for x in rownums_per_items]) + '), 2)'
        else:
            percentage_am_eq = 0
            percentage_am_eq_sp = 0
        if rownums_nonper_items:
            nonpercentage_am_eq = '=ROUND((' + '+'.join(['G' + str(x) for x in rownums_nonper_items]) + '), 2)'
            nonpercentage_am_eq_sp = '=ROUND((' + '+'.join(['H' + str(x) for x in rownums_nonper_items]) + '), 2)'   
        else:
            nonpercentage_am_eq = 0
            nonpercentage_am_eq_sp = 0
        sheet.cell(row=1+row_item, column=7).value = percentage_am_eq
        sheet.cell(row=2+row_item, column=7).value = nonpercentage_am_eq
        sheet.cell(row=3+row_item, column=7).value = '=G' + str(row_item+1) + '+G' + str(row_item+2)  # Up to date amount
        sheet.cell(row=4+row_item, column=7).value = '=ROUND(G' + str(row_item+1) + '*' + str(percentage/100) + ', 2)'  # Percentage amount
        sheet.cell(row=5+row_item, column=7).value = '=G' + str(row_item+3) + '+G' + str(row_item+4)  # Net up to date amount
        if self.prev_bill != None:
            sheet.cell(row=6+row_item, column=7).value = CurrencyR(self.prev_bill.bill_nettotal_amount)  # Previous bill amount
        sheet.cell(row=7+row_item, column=7).value = '=ROUND(G' + str(row_item+5) + '-G' + str(row_item+6) + ', 2)'  # Since previous amount
        sheet.cell(row=8+row_item, column=7).value = '=ROUND(G' + str(row_item+7) + ', 0)'  # Since previous amount rounded
        
        sheet.cell(row=1+row_item, column=8).value = percentage_am_eq_sp
        sheet.cell(row=2+row_item, column=8).value = nonpercentage_am_eq_sp
        sheet.cell(row=3+row_item, column=8).value = '=SUM(H8:H' + str(row_item) + ')'  # Up to date amount
        sheet.cell(row=4+row_item, column=8).value = '=ROUND(H' + str(row_item+1) + '*' + str(percentage/100) + ', 2)'  # Percentage amount
        sheet.cell(row=5+row_item, column=8).value = '=H' + str(row_item+3) + '+H' + str(row_item+4)  # Net up to date amount
        sheet.cell(row=7+row_item, column=8).value = '=ROUND(H' + str(row_item+5) + '-H' + str(row_item+6) + ', 2)'  # Since previous amount
        sheet.cell(row=8+row_item, column=8).value = '=G' + str(row_item+8)  # Since previous amount rounded
        
        sheet.cell(row=4+row_item, column=2).value = sheet.cell(row=4+row_item, column=2).value + ' @ ' + str(percentage) + '%'  # Add percentage value
        
        row_item_sp = row_item + 8
        row_item += 10
        
        # Add details of Adjustments
        for description, amount in self.data.adjustments:
            if amount != 0:
                sheet['B' + str(row_item)] = description
                sheet['G' + str(row_item)] = Currency(amount)
                row_item += 1
        # Add net amount
        if self.data.adjustments:
            sheet['B' + str(row_item)] = 'NET AMOUNT PAYABLE'
            sheet['G' + str(row_item)] = '=ROUND(SUM(G' + str(row_item_sp) + ':G' + str(row_item-1) + '), 0)'
        
        # Bill formatings
        for column in range(1,colend+1):
            # copy coumn widths
            sheet.column_dimensions[get_column_letter(column)].width = \
                template_start_sheet.column_dimensions[get_column_letter(column)].width
        sheet.page_setup.orientation = sheet.ORIENTATION_PORTRAIT
        sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
        sheet.page_setup.fitToHeight = 99
        sheet.page_setup.fitToWidth = 1
        sheet.print_options.horizontalCentered = True
        
        
        ## SHEET 5 PR
        sheet = spreadsheet.create_sheet()
        sheet.title = 'PR'
        template = load_workbook(filename = misc.abs_path('ods_templates','pr.xlsx'))
        rowend = 3
        colend = 7
        rowend_end = 4
        
        # Copy all from PR start
        template_start_sheet = template['start']
        for row in range(1,rowend+1):
            for column in range(1,colend+1):
                sheet.cell(row=row, column=column).value = template_start_sheet.cell(row=row, column=column).value
                sheet.cell(row=row, column=column).font = copy.copy(template_start_sheet.cell(row=row, column=column).font)
                sheet.cell(row=row, column=column).border = copy.copy(template_start_sheet.cell(row=row, column=column).border)
                sheet.cell(row=row, column=column).alignment = copy.copy(template_start_sheet.cell(row=row, column=column).alignment)
        
        # Copy PR items
        
        itemnos = schedule.get_itemnos()
        row_item = rowend+2
        for itemno in itemnos:
            item = schedule[itemno]
            # If item measured, include in bill
            if itemno in self.item_qty and self.item_qty[itemno]:
                # Setup required values
                qty_items = self.item_qty[itemno]
                cmb_refs = self.item_cmb_ref[itemno]
                item_paths = self.item_paths[itemno]
                
                if self.prev_bill is not None and itemno in self.prev_bill.item_normal_amount:
                    sprev_item_normal_amount = self.item_normal_amount[itemno] \
                                               - Decimal(self.prev_bill.item_normal_amount[itemno])
                    sprev_item_excess_amount = self.item_excess_amount[itemno] \
                                               - Decimal(self.prev_bill.item_excess_amount[itemno])
                else:
                    sprev_item_normal_amount = self.item_normal_amount[itemno]
                    sprev_item_excess_amount = self.item_excess_amount[itemno]
                    
                # Copy data
                sheet['A' + str(row_item)] = itemno
                sheet['B' + str(row_item)] = item.extended_description
                sheet['A' + str(row_item)].alignment = Alignment(horizontal='center')
                sheet['B' + str(row_item)].alignment = Alignment(wrap_text=True)
                        
                if self.item_excess_qty[itemno] > 0:
                    row_item += 1
                    sheet['B' + str(row_item)] = 'Qty upto deviation limit of ' + str(item.excess_rate_percent) + '%'
                    sheet['C' + str(row_item)] = item.unit
                    sheet['D' + str(row_item)] = item.rate
                    sheet['E' + str(row_item)] = Currency(item.rate*self.data.item_part_percentage[itemno]/100)
                    sheet['F' + str(row_item)] = self.data.item_part_percentage[itemno]
                    row_item += 1
                    sheet['B' + str(row_item)] = 'Qty above deviation limit of ' + str(item.excess_rate_percent) + '%'
                    sheet['C' + str(row_item)] = item.unit
                    sheet['D' + str(row_item)] = self.data.item_excess_rates[itemno]
                    sheet['E' + str(row_item)] = Currency(self.data.item_excess_rates[itemno]*self.data.item_excess_part_percentage[itemno]/100)
                    sheet['F' + str(row_item)] = self.data.item_excess_part_percentage[itemno]
                    row_item += 2
                else:
                    sheet['C' + str(row_item)] = item.unit
                    sheet['D' + str(row_item)] = item.rate
                    sheet['E' + str(row_item)] = Currency(item.rate*self.data.item_part_percentage[itemno]/100)
                    sheet['F' + str(row_item)] = self.data.item_part_percentage[itemno]
                    row_item += 1
        
        # PR formatings
        for column in range(1,colend+1):
            # copy coumn widths
            sheet.column_dimensions[get_column_letter(column)].width = \
                template_start_sheet.column_dimensions[get_column_letter(column)].width
        sheet.page_setup.orientation = sheet.ORIENTATION_PORTRAIT
        sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
        sheet.page_setup.fitToHeight = 99
        sheet.page_setup.fitToWidth = 1
        sheet.print_options.horizontalCentered = True
        
        
        ## Save Document
        spreadsheet.save(filename)

    def get_text(self):
        total = [str(self.bill_nettotal_amount), str(self.bill_since_prev_amount)]
        if self.data.bill_type == misc.BILL_NORMAL:
            return '<b>' + misc.clean_markup(self.data.title) + '</b> | CMB.No.<b>' + misc.clean_markup(
                self.data.cmb_name) + ' dated ' + misc.clean_markup(self.data.bill_date) + '</b> | TOTAL: <b>' + str(
                total) + '</b>'
        if self.data.bill_type == misc.BILL_FINAL:
            return '<span foreground="green"><b>' + misc.clean_markup(self.data.title) + '</b> | CMB.No.<b>' + misc.clean_markup(
                self.data.cmb_name) + ' dated ' + misc.clean_markup(self.data.bill_date) + '</b> | TOTAL: <b>' + str(
                total) + '</b></span>'
        if self.data.bill_type == misc.BILL_CUSTOM:
            return '<span foreground="red"><b>' + misc.clean_markup(self.data.title) + '</b> | CMB.No.<b>' + misc.clean_markup(
                self.data.cmb_name) + ' dated ' + misc.clean_markup(self.data.bill_date) + '</b> | TOTAL: <b>' + str(
                total) + '</b></span>'

    def print_item(self):
        print("bill " + self.data.title + " start")
        for itemno in self.item_normal_amount:
            if self.data.bill_type in (misc.BILL_NORMAL, misc.BILL_FINAL):
                rate = [self.data.item_excess_rates[count]]
                percentages = [self.data.item_part_percentage[itemno], self.data.item_excess_part_percentage[itemno]]
            elif self.data.bill_type == misc.BILL_CUSTOM:
                rate = []
                percentages = []
            qty = self.data.item_qty[itemno]
            amount = [self.item_normal_amount[itemno], self.item_excess_amount[itemno]]
            print([itemno, qty, rate, percentages, amount])
        print(['\nTotal Amount | Since Prev ', [self.bill_total_amount, self.bill_since_prev_amount]])
        print("bill end")
        
